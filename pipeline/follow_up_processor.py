"""Follow-up worksheet processor.

Normalizes data from the daily follow-up Excel workbook, enriches that data with
portal/channel context from the Accounts directory, and emits structured
``TeamsEvent`` payloads ready for Adaptive Card generation and Teams delivery.

This module bridges the spreadsheet-driven operations workflow with the Teams
notification stack so overdue batches, pending resubmissions, and large
rejections automatically surface with rich context.
"""

from __future__ import annotations

import math
import re
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from integrations.teams.models import (
    EventType,
    StakeholderGroup,
    TeamsEvent,
    TeamsPriority,
)

# ---------------------------------------------------------------------------
# Reference data & helpers
# ---------------------------------------------------------------------------

BRANCH_ALIASES = {
    "riyad": "riyadh",
    "jazan": "jizan",
    "madina": "madinah",
    "medina": "madinah",
    "medinah": "madinah",
    "khamis_mushait": "khamis",
    "onizah": "unizah",
    "onaiza": "unizah",
    "onaizah": "unizah",
}

BRANCH_DISPLAY = {
    "riyadh": "Riyadh",
    "jizan": "Jizan",
    "madinah": "Madinah",
    "khamis": "Khamis Mushait",
    "unizah": "Unaizah",
    "abha": "Abha",
    "makkah": "Makkah",
}

STATUS_ALIASES = {
    "submitted": "submitted",
    "submited": "submitted",
    "submitted ": "submitted",
    "submitted": "submitted",
    "submitted": "submitted",
    "no rejection": "no_rejection",
    "no_rejection": "no_rejection",
    "passed due": "passed_due",
    "passed due ": "passed_due",
    "ready to work": "ready_to_work",
    "ready for work": "ready_to_work",
    "under processing": "under_processing",
    "underprocess": "under_processing",
    "not submitted": "not_submitted",
    "not submit": "not_submitted",
    "submitted-": "submitted",
}

STATUS_DISPLAY = {
    "submitted": "Submitted",
    "no_rejection": "No Rejection",
    "passed_due": "Passed Due",
    "ready_to_work": "Ready To Work",
    "under_processing": "Under Processing",
    "not_submitted": "Not Submitted",
    "unknown": "Needs Review",
}

SPECIAL_HEADER_MAP = {
    "Initial Rejection %": "initial_rejection_percent",
    "Initial Rejection % ": "initial_rejection_percent",
    "Initial Rejected Amount": "initial_rejected_amount",
    "Final Rejection %": "final_rejection_percent",
    "Final Rejection % ": "final_rejection_percent",
    "Due date ": "due_date",
    "Due date": "due_date",
    "Re-submission date ": "resubmission_date",
    "Re-submission date": "resubmission_date",
    "Column1": None,
    "Column44": None,
}

DUE_SOON_THRESHOLD_DAYS = 2
CRITICAL_OVERDUE_THRESHOLD_DAYS = 3
HIGH_REJECTION_AMOUNT = 250_000.0
MEDIUM_REJECTION_AMOUNT = 100_000.0
COMPLIANCE_PERCENT_THRESHOLD = 0.05


def normalize_branch_name(value: Any) -> Optional[str]:
    """Convert raw branch cell content to a canonical slug."""

    if value is None:
        return None
    if isinstance(value, str):
        token = re.sub(r"[^a-zA-Z]", "", value).lower()
    else:
        token = str(value).strip().lower()
    if not token:
        return None
    if token in BRANCH_ALIASES:
        return BRANCH_ALIASES[token]
    if token in BRANCH_DISPLAY:
        return token
    return None


def branch_display_name(branch: Optional[str]) -> str:
    """Return a human-readable branch label."""

    if not branch:
        return "Unknown Branch"
    return BRANCH_DISPLAY.get(branch, branch.replace("_", " ").title())


def slugify(value: Optional[str]) -> str:
    """Create a safe slug for correlation identifiers."""

    if not value:
        return ""
    token = re.sub(r"[^a-zA-Z0-9]+", "-", str(value).strip()).strip("-")
    return token.lower()


def clean_string(value: Any) -> Optional[str]:
    """Normalize string values; treat blanks and placeholders as missing."""

    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned in {"-", "--", "—", "_"}:
            return None
        return cleaned
    return str(value).strip() or None


def parse_float(value: Any) -> Optional[float]:
    """Best-effort numeric parsing with formula/placeholder handling."""

    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned in {"-", "--"}:
            return None
        if cleaned.startswith("="):
            return None
        cleaned = cleaned.replace(",", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def parse_date(value: Any) -> Optional[date]:
    """Parse Excel date representations into ``date`` objects."""

    if value is None:
        return None
    if isinstance(value, datetime):
        candidate = value.date()
    elif isinstance(value, date):
        candidate = value
    elif isinstance(value, (int, float)):
        if value <= 0:
            return None
        # Excel serial dates start at 1899-12-30
        epoch = datetime(1899, 12, 30)
        try:
            candidate = (epoch + timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return None
    elif isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned in {"-", "--", ""}:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
        return None
    else:
        return None

    if candidate.year < 1905:
        # Guard against bogus Excel defaults such as 1900-01-22
        return None
    return candidate


def format_currency(value: Optional[float]) -> str:
    if value is None:
        return "—"
    return f"SAR {value:,.2f}"


def format_percent(value: Optional[float]) -> str:
    if value is None:
        return "—"
    if value <= 1:
        return f"{value * 100:.1f}%"
    return f"{value:.1f}%"


def format_date(value: Optional[date]) -> str:
    if value is None:
        return "Not provided"
    return value.isoformat()


def normalize_status(value: Optional[str]) -> str:
    if not value:
        return "unknown"
    text = re.sub(r"\s+", " ", value.strip().lower())
    if not text:
        return "unknown"
    if text in STATUS_ALIASES:
        return STATUS_ALIASES[text]
    if "pass" in text and "due" in text:
        return "passed_due"
    if "ready" in text:
        return "ready_to_work"
    if "not" in text and "submit" in text:
        return "not_submitted"
    if "under" in text and "process" in text:
        return "under_processing"
    if "submit" in text:
        return "submitted"
    if "no" in text and "rejection" in text:
        return "no_rejection"
    return "unknown"


# ---------------------------------------------------------------------------
# Portal directory ingestion
# ---------------------------------------------------------------------------

class PortalDirectory:
    """Load portal/channel context from the Accounts workbook."""

    def __init__(self, workbook_path: Optional[Path]):
        self._resources: Dict[str, List[Dict[str, str]]] = defaultdict(list)
        self._resource_keys: Dict[str, set] = defaultdict(set)
        self._add_resource(
            "all",
            name="Accounts Workbook",
            description="Refer to Accounts.xlsx for the complete credential and portal mapping.",
        )
        if workbook_path and workbook_path.exists():
            self._load(workbook_path)

    # Public API ---------------------------------------------------------

    def get_resources(self, branch: Optional[str]) -> List[Dict[str, str]]:
        key = branch or "all"
        results: List[Dict[str, str]] = []
        if key in self._resources:
            results.extend(self._resources[key])
        if "all" in self._resources and key != "all":
            results.extend(self._resources["all"])
        return [dict(item) for item in results]

    # Internal helpers ---------------------------------------------------

    def _load(self, workbook_path: Path) -> None:
        wb = load_workbook(workbook_path, data_only=True)
        try:
            if "MOH" in wb.sheetnames:
                self._parse_portal_sheet(wb["MOH"], portal_name="MOH Claim Portal")
            if "Oasis" in wb.sheetnames:
                self._parse_portal_sheet(wb["Oasis"], portal_name="Oasis Portal")
            if "jisr" in wb.sheetnames:
                self._parse_credentials_sheet(
                    wb["jisr"],
                    portal_name="Jisr Workforce Portal",
                    description_prefix="User",
                )
            if "Bupa" in wb.sheetnames:
                self._parse_credentials_sheet(
                    wb["Bupa"],
                    portal_name="Bupa Claims Portal",
                    description_prefix="Account",
                )
            if "Remote" in wb.sheetnames:
                self._parse_remote_sheet(wb["Remote"])
        finally:
            wb.close()

    def _parse_portal_sheet(self, sheet, portal_name: str) -> None:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            branches = {
                normalize_branch_name(cell)
                for cell in row
                if normalize_branch_name(cell)
            }
            urls = [cell for cell in row if isinstance(cell, str) and cell.strip().lower().startswith("http")]
            if not urls:
                continue
            targets = branches or {"all"}
            for url in urls:
                for branch in targets:
                    self._add_resource(
                        branch,
                        name=portal_name,
                        url=url.strip(),
                        description="Portal link",
                    )

    def _parse_credentials_sheet(self, sheet, portal_name: str, description_prefix: str) -> None:
        for branch_cell, username, *_ in sheet.iter_rows(min_row=2, values_only=True):
            branch = normalize_branch_name(branch_cell)
            if not branch:
                continue
            username_hint = self._mask_value(username)
            description = f"{description_prefix} hint: {username_hint}. Credentials stored securely."
            self._add_resource(branch, name=portal_name, description=description)

    def _parse_remote_sheet(self, sheet) -> None:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return
        headers = rows[0]
        for col_index, header in enumerate(headers):
            branch = normalize_branch_name(header)
            if not branch:
                continue
            ip_value = rows[1][col_index] if len(rows) > 1 else None
            ip = clean_string(ip_value)
            if not ip:
                continue
            description = f"Remote desktop IP: {ip}"
            self._add_resource(branch, name="Remote Access", description=description)

    def _add_resource(
        self,
        branch: str,
        name: str,
        url: Optional[str] = None,
        description: Optional[str] = None,
    ) -> None:
        entry_key = (name.lower(), (url or "").lower(), (description or "").lower())
        if entry_key in self._resource_keys[branch]:
            return
        self._resource_keys[branch].add(entry_key)
        resource = {"name": name}
        if url:
            resource["url"] = url
        if description:
            resource["description"] = description
        self._resources[branch].append(resource)

    @staticmethod
    def _mask_value(value: Optional[str]) -> str:
        if not value:
            return "***"
        stripped = value.strip()
        if len(stripped) <= 3:
            return stripped + "***"
        return f"{stripped[:3]}***"


# ---------------------------------------------------------------------------
# Worksheet processor
# ---------------------------------------------------------------------------

class FollowUpWorksheetProcessor:
    """Generate Teams events from the daily follow-up worksheet."""

    def __init__(
        self,
        follow_up_path: Path,
        accounts_path: Optional[Path] = None,
        today: Optional[date] = None,
    ) -> None:
        self.follow_up_path = Path(follow_up_path)
        self.accounts_path = Path(accounts_path) if accounts_path else None
        self.today = today or datetime.utcnow().date()
        self.portal_directory = PortalDirectory(self.accounts_path)

    # Public API ---------------------------------------------------------

    def generate_events(self) -> List[TeamsEvent]:
        """Parse the worksheet and produce actionable Teams events."""

        contexts = self.collect_contexts(include_non_alerts=False)
        return [self._build_event(context) for context in contexts]

    def collect_contexts(self, include_non_alerts: bool = False) -> List[Dict[str, Any]]:
        """Return per-row alert contexts for API consumption and analytics."""

        if not self.follow_up_path.exists():
            raise FileNotFoundError(f"Follow-up workbook not found: {self.follow_up_path}")

        rows = self._load_workbook_rows()
        contexts: List[Dict[str, Any]] = []
        for row in rows:
            normalized = self._normalize_row(row)
            if not normalized:
                continue
            context = self._build_alert_context(normalized)
            if not include_non_alerts and not context["should_alert"]:
                continue
            contexts.append(context)
        return contexts

    # Internal helpers ---------------------------------------------------

    def _load_workbook_rows(self) -> List[Dict[str, Any]]:
        wb = load_workbook(self.follow_up_path, data_only=True)
        try:
            sheet = wb.active
            headers = self._sanitize_headers(sheet)
            rows: List[Dict[str, Any]] = []
            for excel_row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict: Dict[str, Any] = {}
                empty = True
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = excel_row[idx] if idx < len(excel_row) else None
                    if value not in (None, "", " "):
                        empty = False
                    row_dict[header] = value
                if empty:
                    continue
                rows.append(row_dict)
            return rows
        finally:
            wb.close()

    def _sanitize_headers(self, sheet) -> List[Optional[str]]:
        seen: set = set()
        headers: List[Optional[str]] = []
        for cell in sheet[1]:
            raw = cell.value
            if raw in SPECIAL_HEADER_MAP:
                mapped = SPECIAL_HEADER_MAP[raw]
                if mapped is not None:
                    token = mapped
                    base = token
                    counter = 2
                    while token in seen:
                        token = f"{base}_{counter}"
                        counter += 1
                    seen.add(token)
                    headers.append(token)
                else:
                    headers.append(None)
                continue
            if raw is None:
                headers.append(None)
                continue
            token = re.sub(r"[^a-z0-9]+", "_", str(raw).strip().lower()).strip("_")
            if not token:
                headers.append(None)
                continue
            if token == "billing_month":
                token = "billing_month"
            if token == "batch_status":
                token = "batch_status"
            base = token
            counter = 2
            while token in seen:
                token = f"{base}_{counter}"
                counter += 1
            seen.add(token)
            headers.append(token)
        return headers

    def _normalize_row(self, row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        branch_key = normalize_branch_name(row.get("branch"))
        if not branch_key:
            return None
        status_raw = clean_string(row.get("batch_status"))
        status = normalize_status(status_raw)

        due_date = parse_date(row.get("due_date"))
        received_date = parse_date(row.get("received_date"))
        resubmission_date = parse_date(row.get("resubmission_date"))
        days_to_due = None
        if due_date:
            days_to_due = (due_date - self.today).days

        billing_amount = parse_float(row.get("billing_amount"))
        approved_to_pay = parse_float(row.get("approved_to_pay"))
        final_rejection_amount = parse_float(row.get("final_rejection"))
        final_rejection_percent = parse_float(row.get("final_rejection_percent"))
        recovery_amount = parse_float(row.get("recovery_amount"))

        insurance_company = clean_string(row.get("insurance_company")) or "Unknown"
        batch_no = clean_string(row.get("batch_no"))
        processor = clean_string(row.get("processor"))
        rework_type = clean_string(row.get("rework_type"))
        batch_type = clean_string(row.get("batch_type"))
        billing_month = clean_string(row.get("month"))
        year_value = row.get("year")
        year = None
        if isinstance(year_value, (int, float)):
            year = int(year_value)
        elif isinstance(year_value, str) and year_value.strip().isdigit():
            year = int(year_value.strip())

        data = {
            "branch_key": branch_key,
            "branch": branch_display_name(branch_key),
            "status": status,
            "status_display": STATUS_DISPLAY.get(status, status.replace("_", " ").title()),
            "status_raw": status_raw,
            "due_date": due_date,
            "received_date": received_date,
            "resubmission_date": resubmission_date,
            "days_to_due": days_to_due,
            "billing_amount": billing_amount,
            "approved_to_pay": approved_to_pay,
            "final_rejection_amount": final_rejection_amount,
            "final_rejection_percent": final_rejection_percent,
            "recovery_amount": recovery_amount,
            "insurance_company": insurance_company,
            "batch_no": batch_no,
            "processor": processor,
            "rework_type": rework_type,
            "batch_type": batch_type,
            "billing_month": billing_month.title() if billing_month else None,
            "year": year,
        }
        return data

    def _build_alert_context(self, row: Dict[str, Any]) -> Dict[str, Any]:
        alerts: List[str] = []
        days_to_due = row["days_to_due"]
        due_date = row["due_date"]
        final_rejection_amount = row["final_rejection_amount"] or 0.0
        final_rejection_percent = row["final_rejection_percent"]
        recovery_amount = row["recovery_amount"] or 0.0
        status = row["status"]

        def add_alert(message: str) -> None:
            if message not in alerts:
                alerts.append(message)

        if status == "passed_due":
            if due_date and days_to_due is not None:
                add_alert(
                    f"Marked Passed Due – overdue by {abs(days_to_due)} day(s) (was due {format_date(due_date)})"
                )
            else:
                add_alert("Marked Passed Due with missing due date – confirm in worksheet")
        if days_to_due is not None:
            if days_to_due < 0 and status != "passed_due":
                add_alert(
                    f"Due date {format_date(due_date)} passed {abs(days_to_due)} day(s) ago"
                )
            elif 0 <= days_to_due <= DUE_SOON_THRESHOLD_DAYS:
                add_alert(f"Due in {days_to_due} day(s) on {format_date(due_date)}")
        if status == "not_submitted":
            add_alert("Batch flagged as not submitted")
        if status == "ready_to_work":
            add_alert("Batch ready for rework – assign processor")
        if not row.get("processor"):
            add_alert("No processor assigned in worksheet")
        if final_rejection_amount > 0:
            add_alert(f"Final rejection total {format_currency(final_rejection_amount)}")
        if final_rejection_percent and final_rejection_percent >= COMPLIANCE_PERCENT_THRESHOLD:
            add_alert(
                f"Rejection ratio {format_percent(final_rejection_percent)} exceeds threshold"
            )
        if recovery_amount > 0:
            add_alert(f"Recovery amount outstanding {format_currency(recovery_amount)}")

        should_alert = bool(alerts)

        priority = TeamsPriority.INFO
        if days_to_due is not None and days_to_due < 0:
            overdue_days = abs(days_to_due)
            if overdue_days >= CRITICAL_OVERDUE_THRESHOLD_DAYS or final_rejection_amount >= HIGH_REJECTION_AMOUNT:
                priority = TeamsPriority.CRITICAL
            else:
                priority = TeamsPriority.HIGH
        elif status == "not_submitted":
            priority = TeamsPriority.HIGH
        elif status == "ready_to_work":
            priority = TeamsPriority.MEDIUM
        elif final_rejection_amount >= HIGH_REJECTION_AMOUNT:
            priority = TeamsPriority.HIGH
        elif final_rejection_amount >= MEDIUM_REJECTION_AMOUNT or any(
            phrase.startswith("Due in") for phrase in alerts
        ):
            priority = TeamsPriority.MEDIUM

        stakeholders: List[StakeholderGroup] = [StakeholderGroup.NPHIES_INTEGRATION]
        if priority in (TeamsPriority.CRITICAL, TeamsPriority.HIGH):
            stakeholders.append(StakeholderGroup.PMO)
        if final_rejection_percent and final_rejection_percent >= COMPLIANCE_PERCENT_THRESHOLD:
            stakeholders.append(StakeholderGroup.COMPLIANCE)
        elif final_rejection_amount >= HIGH_REJECTION_AMOUNT:
            stakeholders.append(StakeholderGroup.COMPLIANCE)
        stakeholders = list(dict.fromkeys(stakeholders))

        portal_resources = self.portal_directory.get_resources(row["branch_key"])
        worksheet_reference = {
            "name": "Daily Follow-up Worksheet",
            "description": self._worksheet_description(row),
        }
        portals = [worksheet_reference]
        for resource in portal_resources:
            if resource not in portals:
                portals.append(resource)

        correlation_id = row.get("batch_no") or self._fallback_correlation_id(row)

        data_payload = {
            "branch": row["branch"],
            "status_display": row["status_display"],
            "status_raw": row.get("status_raw") or "",
            "insurance_company": row["insurance_company"],
            "batch_no": row.get("batch_no"),
            "processor": row.get("processor"),
            "rework_type": row.get("rework_type"),
            "batch_type": row.get("batch_type"),
            "billing_month": row.get("billing_month"),
            "billing_year": row.get("year"),
            "due_date_display": format_date(row["due_date"]),
            "received_date_display": format_date(row.get("received_date")),
            "resubmission_date_display": format_date(row.get("resubmission_date")),
            "billing_amount_display": format_currency(row.get("billing_amount")),
            "approved_to_pay_display": format_currency(row.get("approved_to_pay")),
            "final_rejection_display": format_currency(row.get("final_rejection_amount")),
            "final_rejection_percent_display": format_percent(row.get("final_rejection_percent")),
            "recovery_amount_display": format_currency(row.get("recovery_amount")),
            "alerts": alerts,
            "portal_resources": portals,
            "days_until_due": days_to_due,
        }

        context = {
            "should_alert": should_alert,
            "priority": priority,
            "stakeholders": stakeholders,
            "correlation_id": correlation_id,
            "data": data_payload,
            "branch": row["branch"],
            "status": status,
            "normalized": dict(row),
        }
        return context

    def _build_event(self, context: Dict[str, Any]) -> TeamsEvent:
        return TeamsEvent(
            event_type=EventType.FOLLOW_UP_STATUS,
            correlation_id=context["correlation_id"],
            priority=context["priority"],
            stakeholders=context["stakeholders"],
            data=context["data"],
        )

    def _fallback_correlation_id(self, row: Dict[str, Any]) -> str:
        branch_slug = slugify(row.get("branch_key")) or "branch"
        payer_slug = slugify(row.get("insurance_company")) or "payer"
        month_slug = slugify(row.get("billing_month")) or "month"
        year = row.get("year") or self.today.year
        return f"followup-{branch_slug}-{payer_slug}-{year}-{month_slug}-{uuid.uuid4().hex[:6]}"

    @staticmethod
    def _worksheet_description(row: Dict[str, Any]) -> str:
        pieces = []
        if row.get("billing_month"):
            pieces.append(row["billing_month"])
        if row.get("year"):
            pieces.append(str(row["year"]))
        if row.get("insurance_company"):
            pieces.append(row["insurance_company"])
        if row.get("batch_no"):
            pieces.append(f"Batch {row['batch_no']}")
        if not pieces:
            return "Worksheet reference"
        return " · ".join(pieces)
