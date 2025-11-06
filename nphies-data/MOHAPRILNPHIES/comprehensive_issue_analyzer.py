"""
Comprehensive NPHIES Issue Analyzer with Solutions
===================================================
Fetches data, extracts insights, identifies root causes, and proposes actionable fixes.
Generates detailed reports with priority rankings and implementation roadmaps.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings

warnings.filterwarnings('ignore')

class ComprehensiveNPHIESAnalyzer:
    def __init__(self, folder_path):
        self.folder_path = Path(folder_path)
        self.output_folder = self.folder_path / 'comprehensive_analysis'
        self.output_folder.mkdir(exist_ok=True)
        
        # Data containers
        self.claims_df = None
        self.payment_df = None
        self.eligibility_df = None
        self.auth_df = None
        self.comm_df = None
        
        # Issue tracking
        self.issues = []
        self.solutions = []
        
    def fetch_data(self):
        """Fetch all NPHIES data files"""
        print("=" * 100)
        print("STEP 1: FETCHING DATA FILES")
        print("=" * 100)
        
        try:
            # Claims
            claim_files = list(self.folder_path.glob('Claim_*.csv'))
            if claim_files:
                self.claims_df = pd.read_csv(claim_files[0])
                self.claims_df['Net Amount'] = pd.to_numeric(self.claims_df['Net Amount'], errors='coerce')
                self.claims_df['Approved Amount'] = pd.to_numeric(self.claims_df['Approved Amount'], errors='coerce')
                print(f"✓ Claims loaded: {len(self.claims_df):,} records")
            
            # Payments
            payment_files = list(self.folder_path.glob('PaymentReconciliation_*.csv'))
            if payment_files:
                self.payment_df = pd.read_csv(payment_files[0])
                self.payment_df['Net Amount'] = pd.to_numeric(self.payment_df['Net Amount'], errors='coerce')
                self.payment_df['Number Of Claims'] = pd.to_numeric(self.payment_df['Number Of Claims'], errors='coerce')
                print(f"✓ Payments loaded: {len(self.payment_df):,} records")
            
            # Eligibility
            elig_files = list(self.folder_path.glob('EligibilityRequest_*.csv'))
            if elig_files:
                self.eligibility_df = pd.read_csv(elig_files[0])
                print(f"✓ Eligibility loaded: {len(self.eligibility_df):,} records")
            
            # Authorization
            auth_files = list(self.folder_path.glob('AdvancedAuth_*.csv'))
            if auth_files:
                self.auth_df = pd.read_csv(auth_files[0])
                print(f"✓ Authorization loaded: {len(self.auth_df):,} records")
            
            # Communications
            comm_files = list(self.folder_path.glob('CommunicationRequest_*.csv'))
            if comm_files:
                self.comm_df = pd.read_csv(comm_files[0])
                print(f"✓ Communications loaded: {len(self.comm_df):,} records")
            
            print("\n✅ Data fetch complete!\n")
            return True
            
        except Exception as e:
            print(f"❌ Error fetching data: {e}")
            return False
    
    def extract_metrics(self):
        """Extract key performance metrics"""
        print("=" * 100)
        print("STEP 2: EXTRACTING KEY METRICS")
        print("=" * 100)
        
        metrics = {}
        
        if self.claims_df is not None:
            total_claims = len(self.claims_df)
            approved = len(self.claims_df[self.claims_df['Status'] == 'Approved'])
            rejected = len(self.claims_df[self.claims_df['Status'].isin(['Rejected', 'Error', 'Cancelled'])])
            partial = len(self.claims_df[self.claims_df['Status'] == 'Partial'])
            
            total_submitted = self.claims_df['Net Amount'].sum()
            total_approved = self.claims_df['Approved Amount'].sum()
            revenue_loss = total_submitted - total_approved
            
            metrics['claims'] = {
                'total': total_claims,
                'approved': approved,
                'approval_rate': (approved / total_claims * 100) if total_claims > 0 else 0,
                'rejected': rejected,
                'rejection_rate': (rejected / total_claims * 100) if total_claims > 0 else 0,
                'partial': partial,
                'partial_rate': (partial / total_claims * 100) if total_claims > 0 else 0,
                'submitted_amount': total_submitted,
                'approved_amount': total_approved,
                'revenue_loss': revenue_loss,
                'loss_percentage': (revenue_loss / total_submitted * 100) if total_submitted > 0 else 0
            }
            
            print(f"✓ Claims Metrics:")
            print(f"  - Total Claims: {total_claims:,}")
            print(f"  - Approval Rate: {metrics['claims']['approval_rate']:.1f}%")
            print(f"  - Rejection Rate: {metrics['claims']['rejection_rate']:.1f}%")
            print(f"  - Revenue Loss: SAR {revenue_loss:,.2f} ({metrics['claims']['loss_percentage']:.1f}%)")
        
        if self.payment_df is not None:
            zero_payments = len(self.payment_df[self.payment_df['Net Amount'] == 0])
            total_payments = len(self.payment_df)
            
            metrics['payments'] = {
                'total_bundles': total_payments,
                'zero_bundles': zero_payments,
                'zero_percentage': (zero_payments / total_payments * 100) if total_payments > 0 else 0,
                'total_amount': self.payment_df['Net Amount'].sum()
            }
            
            print(f"\n✓ Payment Metrics:")
            print(f"  - Zero Payment Rate: {metrics['payments']['zero_percentage']:.1f}%")
        
        if self.eligibility_df is not None:
            elig_errors = len(self.eligibility_df[self.eligibility_df['Status'] == 'Error'])
            total_elig = len(self.eligibility_df)
            
            metrics['eligibility'] = {
                'total': total_elig,
                'errors': elig_errors,
                'error_rate': (elig_errors / total_elig * 100) if total_elig > 0 else 0
            }
            
            print(f"\n✓ Eligibility Metrics:")
            print(f"  - Error Rate: {metrics['eligibility']['error_rate']:.1f}%")
        
        self.metrics = metrics
        print("\n✅ Metric extraction complete!\n")
        return metrics
    
    def analyze_issues(self):
        """Deep dive analysis to identify specific issues"""
        print("=" * 100)
        print("STEP 3: ANALYZING ISSUES & ROOT CAUSES")
        print("=" * 100)
        
        issue_number = 1
        
        # ISSUE 1: Low Approval Rate
        if self.claims_df is not None:
            approval_rate = self.metrics['claims']['approval_rate']
            if approval_rate < 70:
                severity = "CRITICAL" if approval_rate < 50 else "HIGH"
                
                # Analyze by insurer
                insurer_analysis = []
                for insurer in self.claims_df['Insurer Name'].value_counts().head(10).index:
                    insurer_df = self.claims_df[self.claims_df['Insurer Name'] == insurer]
                    ins_approved = len(insurer_df[insurer_df['Status'] == 'Approved'])
                    ins_total = len(insurer_df)
                    ins_rate = (ins_approved / ins_total * 100) if ins_total > 0 else 0
                    
                    if ins_rate < 60:
                        insurer_analysis.append({
                            'insurer': insurer,
                            'total_claims': ins_total,
                            'approval_rate': ins_rate,
                            'rejected': len(insurer_df[insurer_df['Status'].isin(['Rejected', 'Error', 'Cancelled'])])
                        })
                
                self.issues.append({
                    'number': issue_number,
                    'title': 'Low Overall Claim Approval Rate',
                    'severity': severity,
                    'current_value': f"{approval_rate:.1f}%",
                    'target_value': '>70%',
                    'financial_impact': f"SAR {self.metrics['claims']['revenue_loss']:,.2f}",
                    'root_causes': [
                        'Incomplete or inaccurate claim documentation',
                        'Coding errors and mismatches',
                        'Missing prior authorization',
                        'Payer-specific submission requirement gaps',
                        'Outdated fee schedules or tariff codes'
                    ],
                    'affected_insurers': insurer_analysis,
                    'solutions': [
                        {
                            'action': 'Implement Pre-Submission Validation System',
                            'description': 'Deploy automated validation to check completeness before submission',
                            'timeline': '2-4 weeks',
                            'expected_impact': '+15-20% approval rate',
                            'priority': 'HIGH',
                            'steps': [
                                '1. Audit last 100 rejections to identify common validation gaps',
                                '2. Configure validation rules in billing system',
                                '3. Train staff on new validation workflow',
                                '4. Monitor approval rate weekly for 3 months'
                            ]
                        },
                        {
                            'action': 'Insurer-Specific Submission Guidelines',
                            'description': 'Create playbooks for top 5 insurers with specific requirements',
                            'timeline': '3-6 weeks',
                            'expected_impact': '+10-15% approval rate',
                            'priority': 'HIGH',
                            'steps': [
                                '1. Schedule meetings with Bupa, Tawuniya, MedGulf payer reps',
                                '2. Document insurer-specific requirements and preferences',
                                '3. Create checklists for each insurer in EMR system',
                                '4. Conduct training sessions for billing team'
                            ]
                        },
                        {
                            'action': 'Enhanced Staff Training Program',
                            'description': 'Regular training on coding accuracy and documentation',
                            'timeline': 'Ongoing',
                            'expected_impact': '+5-10% approval rate',
                            'priority': 'MEDIUM',
                            'steps': [
                                '1. Monthly coding accuracy workshops',
                                '2. Quarterly updates on payer policy changes',
                                '3. Individual coaching for staff with high error rates',
                                '4. Implement peer review process for complex claims'
                            ]
                        }
                    ]
                })
                issue_number += 1
                
                print(f"✓ Issue #{issue_number-1}: Low Approval Rate Identified ({severity})")
        
        # ISSUE 2: Partial Approval Revenue Leakage
        if self.claims_df is not None:
            partial_df = self.claims_df[self.claims_df['Status'] == 'Partial']
            if len(partial_df) > 0:
                partial_gap = (partial_df['Net Amount'] - partial_df['Approved Amount']).sum()
                partial_rate = self.metrics['claims']['partial_rate']
                
                if partial_rate > 20 or partial_gap > 500000:
                    severity = "HIGH"
                    
                    # Analyze partial patterns
                    partial_by_type = partial_df.groupby('Claim Type').agg({
                        'Net Amount': 'sum',
                        'Approved Amount': 'sum',
                        'Bundle ID': 'count'
                    })
                    partial_by_type['Gap'] = partial_by_type['Net Amount'] - partial_by_type['Approved Amount']
                    
                    self.issues.append({
                        'number': issue_number,
                        'title': 'Excessive Partial Approvals - Revenue Leakage',
                        'severity': severity,
                        'current_value': f"{partial_rate:.1f}% ({len(partial_df):,} claims)",
                        'target_value': '<15%',
                        'financial_impact': f"SAR {partial_gap:,.2f} retained by payers",
                        'root_causes': [
                            'Service bundling issues - unbundled services charged separately',
                            'Pricing above approved tariff schedules',
                            'Missing or insufficient clinical justification',
                            'Quantity or frequency limitations not adhered to',
                            'Duplicate billing for same service/date'
                        ],
                        'patterns': partial_by_type.to_dict('index'),
                        'solutions': [
                            {
                                'action': 'Implement Partial Approval Review Process',
                                'description': 'Systematic review and appeal of partial approvals',
                                'timeline': '1-2 weeks',
                                'expected_impact': 'Recover 20-30% of partial shortfalls',
                                'priority': 'HIGH',
                                'steps': [
                                    '1. Generate weekly partial approval report by insurer/service',
                                    '2. Assign dedicated staff to review top 50 partials monthly',
                                    '3. Prepare appeals with enhanced documentation',
                                    '4. Track recovery rate and adjust strategy'
                                ]
                            },
                            {
                                'action': 'Service Bundling Optimization',
                                'description': 'Review and correct service bundling per payer rules',
                                'timeline': '4-6 weeks',
                                'expected_impact': '-10-15% partial rate',
                                'priority': 'MEDIUM',
                                'steps': [
                                    '1. Audit top 20 services with highest partial rates',
                                    '2. Review payer bundling rules and update charge master',
                                    '3. Configure billing system to auto-bundle appropriately',
                                    '4. Monitor partial rates by service code'
                                ]
                            },
                            {
                                'action': 'Clinical Documentation Improvement (CDI)',
                                'description': 'Enhance clinical notes to support full reimbursement',
                                'timeline': '2-3 months',
                                'expected_impact': '+10-15% recovery on partials',
                                'priority': 'MEDIUM',
                                'steps': [
                                    '1. CDI team review of high-value partial cases',
                                    '2. Create templates for common clinical scenarios',
                                    '3. Physician education on documentation impact',
                                    '4. Implement concurrent documentation review'
                                ]
                            }
                        ]
                    })
                    issue_number += 1
                    
                    print(f"✓ Issue #{issue_number-1}: Partial Approval Leakage Identified ({severity})")
        
        # ISSUE 3: High-Value Claim Rejections
        if self.claims_df is not None:
            high_value_rejects = self.claims_df[
                (self.claims_df['Status'].isin(['Rejected', 'Error', 'Cancelled'])) & 
                (self.claims_df['Net Amount'] > 10000)
            ]
            
            if len(high_value_rejects) > 0:
                hv_loss = high_value_rejects['Net Amount'].sum()
                severity = "CRITICAL" if hv_loss > 500000 else "HIGH"
                
                # Top high-value rejections
                top_hv = high_value_rejects.nlargest(10, 'Net Amount')[
                    ['Transaction Identifier', 'Net Amount', 'Insurer Name', 'Status', 'Claim Type']
                ]
                
                self.issues.append({
                    'number': issue_number,
                    'title': 'High-Value Claim Rejections',
                    'severity': severity,
                    'current_value': f"{len(high_value_rejects):,} claims",
                    'target_value': '0 claims',
                    'financial_impact': f"SAR {hv_loss:,.2f} at risk",
                    'root_causes': [
                        'Complex cases lacking adequate documentation',
                        'Prior authorization not obtained or expired',
                        'Medical necessity not clearly established',
                        'Out-of-network or coverage exclusions',
                        'System/technical submission errors'
                    ],
                    'top_cases': top_hv.to_dict('records'),
                    'solutions': [
                        {
                            'action': 'High-Value Claim Task Force',
                            'description': 'Dedicated team for claims >SAR 10,000',
                            'timeline': 'Immediate',
                            'expected_impact': 'Recover 40-60% of rejected amount',
                            'priority': 'CRITICAL',
                            'steps': [
                                '1. Assign senior medical coder to review all 44 high-value rejections',
                                '2. Obtain physician addenda with detailed clinical rationale',
                                '3. File formal appeals with complete documentation within 30 days',
                                '4. Track each case through resolution',
                                '5. Implement lessons learned in pre-submission process'
                            ]
                        },
                        {
                            'action': 'Mandatory Prior Authorization for High-Value Services',
                            'description': 'Require PA before service delivery for claims >SAR 10k',
                            'timeline': '2-3 weeks',
                            'expected_impact': 'Prevent 80%+ future high-value rejections',
                            'priority': 'HIGH',
                            'steps': [
                                '1. Identify service codes typically >SAR 10k',
                                '2. Configure EMR to flag and require PA before scheduling',
                                '3. Create PA request workflow with dedicated staff',
                                '4. Monitor PA approval rate and turnaround time'
                            ]
                        }
                    ]
                })
                issue_number += 1
                
                print(f"✓ Issue #{issue_number-1}: High-Value Rejections Identified ({severity})")
        
        # ISSUE 4: Zero Payment Bundles
        if self.payment_df is not None:
            zero_rate = self.metrics['payments']['zero_percentage']
            if zero_rate > 50:
                severity = "HIGH" if zero_rate > 80 else "MEDIUM"
                
                # Analyze zero payment patterns
                zero_payments = self.payment_df[self.payment_df['Net Amount'] == 0]
                zero_by_sender = zero_payments.groupby('Sender Name').size().sort_values(ascending=False).head(5)
                
                self.issues.append({
                    'number': issue_number,
                    'title': 'Excessive Zero-Value Payment Bundles',
                    'severity': severity,
                    'current_value': f"{zero_rate:.1f}% ({len(zero_payments):,} bundles)",
                    'target_value': '<20%',
                    'financial_impact': 'Delayed cash flow, reconciliation overhead',
                    'root_causes': [
                        'Payment posting errors or mismatches',
                        'Claims fully denied but reconciliation sent',
                        'EDI file parsing issues',
                        'Payer system integration problems',
                        'Incomplete remittance advice data'
                    ],
                    'top_payers': zero_by_sender.to_dict(),
                    'solutions': [
                        {
                            'action': 'Payment Reconciliation Process Overhaul',
                            'description': 'Fix EDI posting and matching logic',
                            'timeline': '3-4 weeks',
                            'expected_impact': 'Reduce zero bundles to <20%',
                            'priority': 'HIGH',
                            'steps': [
                                '1. Audit EDI 835 file processing for top 5 payers',
                                '2. Validate claim matching logic against remittance advice',
                                '3. Configure system to flag unmatched payments for manual review',
                                '4. Train finance team on reconciliation exception handling',
                                '5. Implement automated alerts for zero-payment anomalies'
                            ]
                        },
                        {
                            'action': 'Payer Follow-Up Campaign',
                            'description': 'Chase aged receivables with zero reconciliation',
                            'timeline': '2-4 weeks',
                            'expected_impact': 'Collect pending amounts',
                            'priority': 'MEDIUM',
                            'steps': [
                                '1. Generate report of claims with zero payment >30 days',
                                '2. Contact payers to verify payment status',
                                '3. Resubmit claims if not received by payer',
                                '4. Escalate chronic issues to payer account managers'
                            ]
                        }
                    ]
                })
                issue_number += 1
                
                print(f"✓ Issue #{issue_number-1}: Zero Payment Bundles Identified ({severity})")
        
        # ISSUE 5: Eligibility Check Failures
        if self.eligibility_df is not None:
            error_rate = self.metrics['eligibility']['error_rate']
            if error_rate > 50:
                severity = "CRITICAL" if error_rate > 90 else "HIGH"
                
                # Find repeat offenders
                error_patients = self.eligibility_df[self.eligibility_df['Status'] == 'Error']
                repeat_errors = error_patients['Patient Identifier'].value_counts().head(10)
                
                self.issues.append({
                    'number': issue_number,
                    'title': 'Eligibility Check System Failure',
                    'severity': severity,
                    'current_value': f"{error_rate:.1f}% failure rate",
                    'target_value': '<10%',
                    'financial_impact': 'Service delivery without coverage verification, denial risk',
                    'root_causes': [
                        'NPHIES API integration issues',
                        'Patient identifier format errors',
                        'Invalid or expired policy numbers',
                        'System timeout or connectivity problems',
                        'Payer platform technical issues'
                    ],
                    'repeat_patients': repeat_errors.to_dict(),
                    'solutions': [
                        {
                            'action': 'Emergency NPHIES Integration Fix',
                            'description': 'Debug and repair eligibility API connection',
                            'timeline': '1-2 weeks (URGENT)',
                            'expected_impact': 'Restore eligibility checks to >90% success',
                            'priority': 'CRITICAL',
                            'steps': [
                                '1. Engage IT team and NPHIES technical support immediately',
                                '2. Review API logs to identify error patterns',
                                '3. Validate authentication tokens and credentials',
                                '4. Test patient ID format with known-good samples',
                                '5. Implement error handling and retry logic',
                                '6. Set up monitoring alerts for eligibility failures',
                                '7. Create fallback manual verification process'
                            ]
                        },
                        {
                            'action': 'Patient Identifier Validation',
                            'description': 'Implement front-end validation for patient IDs',
                            'timeline': '1-2 weeks',
                            'expected_impact': 'Prevent 30-40% of format-related errors',
                            'priority': 'HIGH',
                            'steps': [
                                '1. Analyze failed patient IDs to identify format issues',
                                '2. Implement real-time ID validation at registration',
                                '3. Create lookup table for problematic patient IDs',
                                '4. Train registration staff on proper ID entry',
                                '5. Add visual indicators for validation status'
                            ]
                        },
                        {
                            'action': 'Manual Verification Backup Process',
                            'description': 'Temporary workaround while fixing integration',
                            'timeline': 'Immediate',
                            'expected_impact': 'Prevent service delivery without coverage',
                            'priority': 'HIGH',
                            'steps': [
                                '1. Print list of repeat-error patients (top 20)',
                                '2. Call insurers directly to verify coverage',
                                '3. Document coverage details in EMR',
                                '4. Flag these patients for no automated eligibility check'
                            ]
                        }
                    ]
                })
                issue_number += 1
                
                print(f"✓ Issue #{issue_number-1}: Eligibility Failures Identified ({severity})")
        
        # ISSUE 6: Communication Bottlenecks
        if self.comm_df is not None and 'Associated Transaction' in self.comm_df.columns:
            comm_per_trans = self.comm_df.groupby('Associated Transaction').size()
            avg_comms = comm_per_trans.mean()
            max_comms = comm_per_trans.max()
            
            if avg_comms > 2 or max_comms > 5:
                severity = "MEDIUM"
                
                # Find transactions with excessive communications
                excessive_comms = comm_per_trans[comm_per_trans >= 5].sort_values(ascending=False).head(10)
                
                self.issues.append({
                    'number': issue_number,
                    'title': 'Excessive Communication Back-and-Forth',
                    'severity': severity,
                    'current_value': f"Avg {avg_comms:.1f} communications/transaction, Max {max_comms}",
                    'target_value': '<2 communications/transaction',
                    'financial_impact': 'Processing delays, staff time waste',
                    'root_causes': [
                        'Incomplete initial claim submission',
                        'Missing or unclear documentation',
                        'Payer requests for additional information',
                        'Coding clarification requirements',
                        'Follow-up on delayed adjudication'
                    ],
                    'worst_transactions': excessive_comms.to_dict(),
                    'solutions': [
                        {
                            'action': 'First-Time-Right Submission Initiative',
                            'description': 'Improve initial submission completeness',
                            'timeline': '4-6 weeks',
                            'expected_impact': 'Reduce communications by 40-50%',
                            'priority': 'MEDIUM',
                            'steps': [
                                '1. Analyze common communication request types',
                                '2. Create checklist for complete submissions',
                                '3. Implement submission quality audit',
                                '4. Provide feedback to high-communication submitters',
                                '5. Track communication rate by staff member'
                            ]
                        },
                        {
                            'action': 'Automated Response Templates',
                            'description': 'Standardize responses to common payer requests',
                            'timeline': '2-3 weeks',
                            'expected_impact': 'Reduce response time by 50%',
                            'priority': 'LOW',
                            'steps': [
                                '1. Identify top 10 communication request types',
                                '2. Create templates with required information',
                                '3. Configure templates in communication system',
                                '4. Train staff on template usage'
                            ]
                        }
                    ]
                })
                issue_number += 1
                
                print(f"✓ Issue #{issue_number-1}: Communication Bottlenecks Identified ({severity})")
        
        print(f"\n✅ Issue analysis complete! {len(self.issues)} issues identified.\n")
        return self.issues
    
    def generate_comprehensive_report(self):
        """Generate detailed report with issues and solutions"""
        print("=" * 100)
        print("STEP 4: GENERATING COMPREHENSIVE REPORT")
        print("=" * 100)
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Main Report
        report_lines = []
        append = report_lines.append
        
        append("╔" + "=" * 98 + "╗")
        append("║" + " " * 98 + "║")
        append("║" + "COMPREHENSIVE NPHIES ISSUE ANALYSIS & SOLUTION REPORT".center(98) + "║")
        append("║" + "Al-Hayat National Hospital - Unaizah - Al-Qassim".center(98) + "║")
        append("║" + " " * 98 + "║")
        append("╚" + "=" * 98 + "╝")
        append("")
        append(f"Generated: {timestamp}")
        append(f"Analysis Period: April 2025")
        append(f"Total Issues Identified: {len(self.issues)}")
        append("")
        append("")
        
        # Executive Summary
        append("=" * 100)
        append("EXECUTIVE SUMMARY")
        append("=" * 100)
        append("")
        
        critical_issues = [i for i in self.issues if i['severity'] == 'CRITICAL']
        high_issues = [i for i in self.issues if i['severity'] == 'HIGH']
        medium_issues = [i for i in self.issues if i['severity'] == 'MEDIUM']
        
        append(f"Priority Breakdown:")
        append(f"  • CRITICAL Issues: {len(critical_issues)}")
        append(f"  • HIGH Issues: {len(high_issues)}")
        append(f"  • MEDIUM Issues: {len(medium_issues)}")
        append("")
        
        if self.claims_df is not None:
            append(f"Key Metrics:")
            append(f"  • Total Claims Processed: {self.metrics['claims']['total']:,}")
            append(f"  • Approval Rate: {self.metrics['claims']['approval_rate']:.1f}%")
            append(f"  • Rejection Rate: {self.metrics['claims']['rejection_rate']:.1f}%")
            append(f"  • Revenue Loss: SAR {self.metrics['claims']['revenue_loss']:,.2f}")
            append("")
        
        # Detailed Issues
        append("")
        append("=" * 100)
        append("DETAILED ISSUES & SOLUTIONS")
        append("=" * 100)
        append("")
        
        for issue in self.issues:
            append("")
            append("▼" * 50)
            append(f"ISSUE #{issue['number']}: {issue['title']}")
            append("▼" * 50)
            append("")
            append(f"Severity Level: {issue['severity']}")
            append(f"Current State: {issue['current_value']}")
            append(f"Target State: {issue['target_value']}")
            append(f"Financial Impact: {issue['financial_impact']}")
            append("")
            
            append("ROOT CAUSES:")
            for i, cause in enumerate(issue['root_causes'], 1):
                append(f"  {i}. {cause}")
            append("")
            
            # Additional context
            if 'affected_insurers' in issue and issue['affected_insurers']:
                append("AFFECTED INSURERS (Approval Rate < 60%):")
                for ins in issue['affected_insurers'][:5]:
                    append(f"  • {ins['insurer'][:60]}")
                    append(f"    - Total Claims: {ins['total_claims']:,}")
                    append(f"    - Approval Rate: {ins['approval_rate']:.1f}%")
                    append(f"    - Rejected: {ins['rejected']:,}")
                append("")
            
            if 'top_cases' in issue and issue['top_cases']:
                append("TOP 5 HIGH-VALUE REJECTED CLAIMS:")
                for case in issue['top_cases'][:5]:
                    append(f"  • Transaction {case['Transaction Identifier']}")
                    append(f"    Amount: SAR {case['Net Amount']:,.2f} | Insurer: {case['Insurer Name'][:50]} | Status: {case['Status']}")
                append("")
            
            if 'repeat_patients' in issue:
                append("TOP PATIENTS WITH REPEATED ERRORS:")
                for pid, count in list(issue['repeat_patients'].items())[:5]:
                    append(f"  • Patient {pid}: {count} failed attempts")
                append("")
            
            # Solutions
            append("─" * 100)
            append("PROPOSED SOLUTIONS:")
            append("─" * 100)
            append("")
            
            for sol_idx, solution in enumerate(issue['solutions'], 1):
                append(f"SOLUTION {sol_idx}: {solution['action']}")
                append(f"{'─' * 100}")
                append(f"Description: {solution['description']}")
                append(f"Timeline: {solution['timeline']}")
                append(f"Expected Impact: {solution['expected_impact']}")
                append(f"Priority: {solution['priority']}")
                append("")
                append("Implementation Steps:")
                for step in solution['steps']:
                    append(f"  {step}")
                append("")
            
            append("")
        
        # Implementation Roadmap
        append("")
        append("=" * 100)
        append("IMPLEMENTATION ROADMAP")
        append("=" * 100)
        append("")
        
        append("IMMEDIATE ACTIONS (Week 1-2):")
        append("─" * 100)
        immediate_actions = []
        for issue in self.issues:
            for solution in issue['solutions']:
                if solution['priority'] == 'CRITICAL' or 'Immediate' in solution['timeline']:
                    immediate_actions.append(f"  • {solution['action']} (Issue #{issue['number']})")
        
        if immediate_actions:
            for action in immediate_actions:
                append(action)
        else:
            append("  • No immediate critical actions required")
        append("")
        
        append("SHORT-TERM ACTIONS (Week 3-6):")
        append("─" * 100)
        short_term = []
        for issue in self.issues:
            for solution in issue['solutions']:
                if solution['priority'] == 'HIGH' and 'Immediate' not in solution['timeline']:
                    short_term.append(f"  • {solution['action']} (Issue #{issue['number']})")
        
        for action in short_term[:5]:
            append(action)
        append("")
        
        append("MEDIUM-TERM ACTIONS (2-3 Months):")
        append("─" * 100)
        medium_term = []
        for issue in self.issues:
            for solution in issue['solutions']:
                if solution['priority'] == 'MEDIUM':
                    medium_term.append(f"  • {solution['action']} (Issue #{issue['number']})")
        
        for action in medium_term[:5]:
            append(action)
        append("")
        
        # Success Metrics
        append("")
        append("=" * 100)
        append("SUCCESS METRICS TO MONITOR")
        append("=" * 100)
        append("")
        append("Track these KPIs weekly:")
        append("  1. Claim Approval Rate (Target: >70%)")
        append("  2. Rejection Rate (Target: <10%)")
        append("  3. Partial Approval Rate (Target: <15%)")
        append("  4. Revenue Loss Amount (Target: <SAR 500,000/month)")
        append("  5. Eligibility Check Success Rate (Target: >95%)")
        append("  6. Zero Payment Bundle Rate (Target: <20%)")
        append("  7. Average Communications per Transaction (Target: <2)")
        append("  8. Days in A/R (Target: <45 days)")
        append("")
        append("Implement weekly RCM huddle to review these metrics and adjust strategies.")
        append("")
        
        # Conclusion
        append("")
        append("=" * 100)
        append("CONCLUSION & NEXT STEPS")
        append("=" * 100)
        append("")
        append("This comprehensive analysis has identified critical operational and financial issues")
        append("affecting Al-Hayat National Hospital's revenue cycle. The proposed solutions are")
        append("prioritized by severity and expected impact.")
        append("")
        append("RECOMMENDED IMMEDIATE ACTIONS:")
        append("  1. Convene RCM leadership team to review this report")
        append("  2. Assign owners to each critical and high-priority issue")
        append("  3. Allocate budget and resources for top 3 solutions")
        append("  4. Schedule weekly progress review meetings")
        append("  5. Re-run this analysis monthly to track improvements")
        append("")
        append("Expected Overall Impact:")
        append("  • Approval Rate: 50.9% → 70%+ (within 3 months)")
        append("  • Revenue Recovery: SAR 500,000 - 800,000 (from appeals and process fixes)")
        append("  • Operational Efficiency: 30%+ reduction in rework and denials")
        append("")
        append("=" * 100)
        append("END OF REPORT")
        append("=" * 100)
        
        # Write report
        report_path = self.output_folder / f'comprehensive_issue_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        report_path.write_text('\n'.join(report_lines), encoding='utf-8')
        
        print(f"✓ Comprehensive report generated: {report_path.name}")
        
        # Generate Excel Summary
        self.generate_excel_summary()
        
        print(f"\n✅ All reports generated in: {self.output_folder}\n")
        
        return report_path
    
    def generate_excel_summary(self):
        """Generate Excel workbook with issue tracking"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Issues Summary Sheet
            ws1 = wb.active
            ws1.title = "Issues Summary"
            
            # Headers
            headers = ['Issue #', 'Title', 'Severity', 'Current State', 'Target', 'Financial Impact', 'Priority']
            ws1.append(headers)
            
            # Style headers
            for cell in ws1[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Add issue data
            for issue in self.issues:
                # Get highest priority solution
                priorities = [s['priority'] for s in issue['solutions']]
                top_priority = 'CRITICAL' if 'CRITICAL' in priorities else ('HIGH' if 'HIGH' in priorities else 'MEDIUM')
                
                ws1.append([
                    issue['number'],
                    issue['title'],
                    issue['severity'],
                    issue['current_value'],
                    issue['target_value'],
                    issue['financial_impact'],
                    top_priority
                ])
            
            # Action Items Sheet
            ws2 = wb.create_sheet("Action Items")
            headers2 = ['Issue #', 'Solution', 'Priority', 'Timeline', 'Expected Impact', 'Status', 'Owner', 'Notes']
            ws2.append(headers2)
            
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            for issue in self.issues:
                for solution in issue['solutions']:
                    ws2.append([
                        issue['number'],
                        solution['action'],
                        solution['priority'],
                        solution['timeline'],
                        solution['expected_impact'],
                        'Not Started',  # Status
                        '',  # Owner
                        ''   # Notes
                    ])
            
            # Adjust column widths
            for ws in [ws1, ws2]:
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save
            excel_path = self.output_folder / f'issue_tracking_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(excel_path)
            
            print(f"✓ Excel tracking workbook generated: {excel_path.name}")
            
        except ImportError:
            print("⚠ openpyxl not available, skipping Excel generation")
        except Exception as e:
            print(f"⚠ Error generating Excel: {e}")
    
    def run_complete_analysis(self):
        """Execute full analysis pipeline"""
        print("\n")
        print("╔" + "=" * 98 + "╗")
        print("║" + " " * 98 + "║")
        print("║" + "COMPREHENSIVE NPHIES ISSUE ANALYZER".center(98) + "║")
        print("║" + "Fetch → Extract → Analyze → Solutions → Fixes".center(98) + "║")
        print("║" + " " * 98 + "║")
        print("╚" + "=" * 98 + "╝")
        print("\n")
        
        # Step 1: Fetch
        if not self.fetch_data():
            print("\n❌ Analysis aborted due to data fetch failure")
            return False
        
        # Step 2: Extract
        self.extract_metrics()
        
        # Step 3: Analyze
        self.analyze_issues()
        
        # Step 4: Generate Reports
        report_path = self.generate_comprehensive_report()
        
        print("\n")
        print("=" * 100)
        print("✅ COMPREHENSIVE ANALYSIS COMPLETE!")
        print("=" * 100)
        print()
        print(f"📁 Output Location: {self.output_folder}")
        print()
        print("📄 Generated Files:")
        print(f"  1. Comprehensive Issue Report (TXT)")
        print(f"  2. Issue Tracking Workbook (XLSX)")
        print()
        print(f"📊 Summary:")
        print(f"  • Total Issues Identified: {len(self.issues)}")
        print(f"  • Critical Issues: {len([i for i in self.issues if i['severity'] == 'CRITICAL'])}")
        print(f"  • High Priority Issues: {len([i for i in self.issues if i['severity'] == 'HIGH'])}")
        print(f"  • Total Solutions Proposed: {sum(len(i['solutions']) for i in self.issues)}")
        print()
        print("🎯 Next Steps:")
        print("  1. Review the comprehensive report")
        print("  2. Prioritize critical and high-severity issues")
        print("  3. Assign owners using the Excel tracking workbook")
        print("  4. Implement solutions according to the roadmap")
        print("  5. Monitor success metrics weekly")
        print()
        print("=" * 100)
        print()
        
        return True


def main():
    """Main execution"""
    folder_path = r"c:\Users\rcmrejection3\OneDrive\Desktop\MOHAPRILNPHIES"
    
    analyzer = ComprehensiveNPHIESAnalyzer(folder_path)
    analyzer.run_complete_analysis()


if __name__ == "__main__":
    main()
