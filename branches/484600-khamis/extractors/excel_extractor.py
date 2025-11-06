"""
Excel Data Extractor for Bupa Claims Documents
"""
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
from datetime import datetime
import openpyxl


class ExcelExtractor:
    """Extract data from Bupa Excel documents"""
    
    def __init__(self, excel_path: str):
        """
        Initialize Excel extractor
        
        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.sheets_data = {}
        
    def extract_all(self) -> Dict[str, Any]:
        """
        Extract all data from Excel file
        
        Returns:
            Dictionary containing all extracted data
        """
        # Load workbook
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        # Extract data from all sheets
        for sheet_name in self.workbook.sheetnames:
            self.sheets_data[sheet_name] = self._extract_sheet(sheet_name)
        
        return {
            'metadata': self._extract_metadata(),
            'sheets': self.sheets_data,
            'provider_info': self._extract_provider_info(),
            'claims': self._extract_claims(),
            'financial_summary': self._extract_financial_summary(),
        }
    
    def _extract_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Extract data from a specific sheet"""
        df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
        
        return {
            'name': sheet_name,
            'shape': df.shape,
            'columns': df.columns.tolist(),
            'data': df.to_dict('records'),
            'dataframe': df
        }
    
    def _extract_metadata(self) -> Dict[str, Any]:
        """Extract Excel file metadata"""
        properties = self.workbook.properties
        
        return {
            'filename': self.excel_path.name,
            'file_size': self.excel_path.stat().st_size,
            'extracted_at': datetime.now().isoformat(),
            'sheet_names': self.workbook.sheetnames,
            'sheet_count': len(self.workbook.sheetnames),
            'creator': properties.creator if properties else None,
            'created': properties.created.isoformat() if properties and properties.created else None,
            'modified': properties.modified.isoformat() if properties and properties.modified else None,
        }
    
    def _extract_provider_info(self) -> Dict[str, Any]:
        """Extract provider information from Excel"""
        provider_info = {
            'provider_id': '484600',
            'provider_name': '',
            'statement_number': '',
            'statement_date': ''
        }
        
        # Try to find provider info in first sheet
        if self.sheets_data:
            first_sheet = list(self.sheets_data.values())[0]
            df = first_sheet['dataframe']
            
            # Look for common header patterns
            for col in df.columns:
                col_lower = str(col).lower()
                if 'provider' in col_lower and 'id' in col_lower:
                    if not df.empty:
                        provider_info['provider_id'] = str(df[col].iloc[0])
                elif 'provider' in col_lower and 'name' in col_lower:
                    if not df.empty:
                        provider_info['provider_name'] = str(df[col].iloc[0])
                elif 'statement' in col_lower and 'number' in col_lower:
                    if not df.empty:
                        provider_info['statement_number'] = str(df[col].iloc[0])
                elif 'statement' in col_lower and 'date' in col_lower:
                    if not df.empty:
                        provider_info['statement_date'] = str(df[col].iloc[0])
        
        return provider_info
    
    def _extract_claims(self) -> List[Dict[str, Any]]:
        """Extract individual claims from Excel"""
        claims = []
        
        # Find the sheet most likely to contain claims data
        claims_sheet = None
        for sheet_name, sheet_data in self.sheets_data.items():
            # Look for keywords in sheet name
            if any(keyword in sheet_name.lower() for keyword in ['claim', 'detail', 'transaction']):
                claims_sheet = sheet_data
                break
        
        # If no specific claims sheet found, use the largest sheet
        if not claims_sheet and self.sheets_data:
            claims_sheet = max(self.sheets_data.values(), key=lambda x: x['shape'][0])
        
        if claims_sheet and claims_sheet['data']:
            claims = claims_sheet['data']
            
            # Clean up claims data
            cleaned_claims = []
            for claim in claims:
                # Remove None keys and empty values
                cleaned_claim = {k: v for k, v in claim.items() 
                               if k is not None and pd.notna(v)}
                if cleaned_claim:
                    cleaned_claims.append(cleaned_claim)
            
            claims = cleaned_claims
        
        return claims
    
    def _extract_financial_summary(self) -> Dict[str, Any]:
        """Extract financial summary from Excel"""
        summary = {
            'total_claims': 0,
            'total_amount': 0.0,
            'approved_amount': 0.0,
            'rejected_amount': 0.0,
            'pending_amount': 0.0,
            'currency': 'AED'
        }
        
        # Try to calculate from claims data
        if self.sheets_data:
            for sheet_data in self.sheets_data.values():
                df = sheet_data['dataframe']
                
                # Look for amount columns
                amount_columns = [col for col in df.columns 
                                if any(keyword in str(col).lower() 
                                     for keyword in ['amount', 'value', 'total', 'sum'])]
                
                if amount_columns:
                    summary['total_claims'] = len(df)
                    
                    # Sum up amounts
                    for col in amount_columns:
                        try:
                            total = pd.to_numeric(df[col], errors='coerce').sum()
                            if pd.notna(total) and total > 0:
                                col_lower = str(col).lower()
                                if 'approved' in col_lower:
                                    summary['approved_amount'] = float(total)
                                elif 'rejected' in col_lower or 'denied' in col_lower:
                                    summary['rejected_amount'] = float(total)
                                elif 'pending' in col_lower:
                                    summary['pending_amount'] = float(total)
                                else:
                                    summary['total_amount'] = max(summary['total_amount'], float(total))
                        except:
                            continue
        
        return summary
    
    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names"""
        if not self.workbook:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        return self.workbook.sheetnames
    
    def get_sheet_data(self, sheet_name: str) -> pd.DataFrame:
        """Get data from specific sheet as DataFrame"""
        return pd.read_excel(self.excel_path, sheet_name=sheet_name)
    
    def export_to_csv(self, output_dir: str = 'exports'):
        """Export all sheets to CSV files"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        exported_files = []
        for sheet_name in self.get_sheet_names():
            df = self.get_sheet_data(sheet_name)
            csv_filename = f"{self.excel_path.stem}_{sheet_name}.csv"
            csv_path = output_path / csv_filename
            df.to_csv(csv_path, index=False)
            exported_files.append(str(csv_path))
        
        return exported_files


if __name__ == '__main__':
    # Example usage
    import sys
    import json
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        extractor = ExcelExtractor(excel_path)
        data = extractor.extract_all()
        
        print("=== Extracted Data ===")
        print(f"Metadata: {data['metadata']}")
        print(f"Provider Info: {data['provider_info']}")
        print(f"Financial Summary: {data['financial_summary']}")
        print(f"Number of Claims: {len(data['claims'])}")
        print(f"Number of Sheets: {len(data['sheets'])}")
        
        # Show first few claims
        if data['claims']:
            print("\n=== Sample Claims (first 3) ===")
            for i, claim in enumerate(data['claims'][:3], 1):
                print(f"\nClaim {i}:")
                print(json.dumps(claim, indent=2, default=str))
    else:
        print("Usage: python excel_extractor.py <path_to_excel>")
