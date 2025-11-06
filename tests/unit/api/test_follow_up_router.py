from datetime import date, timedelta
from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from config.settings import settings
from routers.follow_up_router import router as follow_up_router


def _build_follow_up_workbook(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Follow up"
    headers = [
        "Month",
        "Branch",
        "Insurance Company",
        "Billing Amount",
        "Approved to pay",
        "Final Rejection",
        "Final Rejection %",
        "Recovery Amount",
        "Batch No.",
        "Rework type",
        "Batch Type",
        "Received Date",
        "Due date",
        "Processor",
        "Batch Status",
        "Re-submission date",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _build_accounts_workbook(path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    moh = wb.create_sheet("MOH")
    moh.append(["Approval Portal", "Riyadh"])
    moh.append([None, "Riyadh", "https://portal.example/claims"])
    wb.save(path)
    wb.close()


def test_follow_up_router_returns_filtered_records(tmp_path, monkeypatch):
    follow_up_path = tmp_path / "daily-follow-ups.xlsx"
    accounts_path = tmp_path / "Accounts.xlsx"

    future_received = date.today()
    future_due = future_received + timedelta(days=10)

    _build_follow_up_workbook(
        follow_up_path,
        [
            [
                "Oct",
                "Riyadh",
                "MOH",
                1_000_000,
                800_000,
                250_000,
                0.2,
                50_000,
                "B123",
                "Re-submission",
                "IP",
                date(2024, 9, 1),
                date(2024, 10, 15),
                "Dr. Mutasim",
                "Passed Due",
                date(2024, 10, 20),
            ],
            [
                "Nov",
                "Jizan",
                "TAWUNIYA",
                900_000,
                900_000,
                0,
                0,
                0,
                "B999",
                "Re-submission",
                "IP",
                future_received,
                future_due,
                "Dr. Ahmed",
                "Submitted",
                future_received + timedelta(days=2),
            ],
        ],
    )
    _build_accounts_workbook(accounts_path)

    # Monkeypatch environment settings
    monkeypatch.setattr(settings, "FOLLOW_UP_WORKBOOK_PATH", str(follow_up_path))
    monkeypatch.setattr(settings, "ACCOUNTS_WORKBOOK_PATH", str(accounts_path))

    app = FastAPI()
    app.include_router(follow_up_router)

    client = TestClient(app)

    response = client.get("/api/follow-ups", params={"priority": "critical"})
    assert response.status_code == 200

    payload = response.json()
    assert payload["summary"]["total_rows"] == 1
    assert payload["summary"]["critical"] == 1
    assert payload["overall_summary"]["total_rows"] == 2
    assert len(payload["records"]) == 1
    assert payload["records"][0]["branch_key"] == "riyadh"