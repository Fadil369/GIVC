from datetime import date
from pathlib import Path

from openpyxl import Workbook

from integrations.teams.models import EventType, StakeholderGroup, TeamsPriority
from pipeline.follow_up_processor import FollowUpWorksheetProcessor


def _build_follow_up_workbook(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Follow up"
    headers = [
        "Month",
        "Branch",
        "Insurance Company",
        "Billing Amount",
        "Approved to pay",
        "Final Rejection",
        "Final Rejection %",
        "Recovery Amount",
        "Batch No.",
        "Rework type",
        "Batch Type",
        "Received Date",
        "Due date",
        "Processor",
        "Batch Status",
        "Re-submission date",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _build_accounts_workbook(path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    moh = wb.create_sheet("MOH")
    moh.append(["Approval Portal", "Riyadh", "Jizan"])
    moh.append([None, "Riyadh", "https://portal.example/claims"])

    remote = wb.create_sheet("Remote")
    remote.append(["Column1", "Riyadh", "Jizan"])
    remote.append(["IP Adress", "10.0.10.10", "10.0.20.20"])
    remote.append(["User Name", "corp\\riyadh", "corp\\jizan"])
    remote.append(["Password", "secret", "secret"])

    wb.save(path)
    wb.close()


def test_generate_events_overdue_batches(tmp_path):
    follow_up_path = tmp_path / "daily-follow-ups.xlsx"
    accounts_path = tmp_path / "Accounts.xlsx"

    _build_follow_up_workbook(
        follow_up_path,
        [
            [
                "Oct",
                "Riyadh",
                "MOH",
                1_000_000,
                800_000,
                250_000,
                0.2,
                50_000,
                "B123",
                "Re-submission",
                "IP",
                date(2024, 9, 1),
                date(2024, 10, 15),
                "Dr. Mutasim",
                "Passed Due",
                date(2024, 10, 20),
            ]
        ],
    )
    _build_accounts_workbook(accounts_path)

    processor = FollowUpWorksheetProcessor(
        follow_up_path=follow_up_path,
        accounts_path=accounts_path,
        today=date(2024, 10, 20),
    )

    events = processor.generate_events()

    assert len(events) == 1
    event = events[0]
    assert event.event_type == EventType.FOLLOW_UP_STATUS
    assert event.priority == TeamsPriority.CRITICAL
    assert StakeholderGroup.NPHIES_INTEGRATION in event.stakeholders
    assert StakeholderGroup.PMO in event.stakeholders
    assert StakeholderGroup.COMPLIANCE in event.stakeholders
    assert any("Passed Due" in alert or "overdue" in alert.lower() for alert in event.data["alerts"])
    portals = event.data["portal_resources"]
    assert any(resource["name"] == "MOH Claim Portal" for resource in portals)
    assert any("Remote" in resource["name"] for resource in portals)


def test_generate_events_skips_non_actionable_rows(tmp_path):
    follow_up_path = tmp_path / "daily-follow-ups.xlsx"

    _build_follow_up_workbook(
        follow_up_path,
        [
            [
                "Nov",
                "Jizan",
                "TAWUNIYA",
                900_000,
                900_000,
                0,
                0,
                0,
                "B999",
                "Re-submission",
                "IP",
                date(2024, 11, 1),
                date(2024, 12, 1),
                "Dr. Ahmed",
                "Submitted",
                date(2024, 11, 10),
            ]
        ],
    )

    processor = FollowUpWorksheetProcessor(
        follow_up_path=follow_up_path,
        accounts_path=None,
        today=date(2024, 11, 15),
    )

    events = processor.generate_events()

    assert events == []
