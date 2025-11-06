from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

from integrations.teams.models import TeamsPriority
from services.follow_up import FollowUpWorksheetService


def _build_follow_up_workbook(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Follow up"
    headers = [
        "Month",
        "Branch",
        "Insurance Company",
        "Billing Amount",
        "Approved to pay",
        "Final Rejection",
        "Final Rejection %",
        "Recovery Amount",
        "Batch No.",
        "Rework type",
        "Batch Type",
        "Received Date",
        "Due date",
        "Processor",
        "Batch Status",
        "Re-submission date",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _build_accounts_workbook(path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    moh = wb.create_sheet("MOH")
    moh.append(["Approval Portal", "Riyadh", "Jizan"])
    moh.append([None, "Riyadh", "https://portal.example/claims"])

    wb.save(path)
    wb.close()


def test_follow_up_snapshot_contains_records_and_summary(tmp_path):
    follow_up_path = tmp_path / "daily-follow-ups.xlsx"
    accounts_path = tmp_path / "Accounts.xlsx"

    future_received = date.today()
    future_due = future_received + timedelta(days=10)

    _build_follow_up_workbook(
        follow_up_path,
        [
            [
                "Oct",
                "Riyadh",
                "MOH",
                1_000_000,
                800_000,
                250_000,
                0.2,
                50_000,
                "B123",
                "Re-submission",
                "IP",
                date(2024, 9, 1),
                date(2024, 10, 15),
                "Dr. Mutasim",
                "Passed Due",
                date(2024, 10, 20),
            ],
            [
                "Nov",
                "Jizan",
                "TAWUNIYA",
                900_000,
                900_000,
                0,
                0,
                0,
                "B999",
                "Re-submission",
                "IP",
                future_received,
                future_due,
                "Dr. Ahmed",
                "Submitted",
                future_received + timedelta(days=2),
            ],
        ],
    )
    _build_accounts_workbook(accounts_path)

    service = FollowUpWorksheetService(follow_up_path=follow_up_path, accounts_path=accounts_path)
    snapshot = service.get_snapshot(include_non_alerts=True)

    assert snapshot.records, "snapshot should contain records"
    assert snapshot.summary.total_rows == 2
    assert snapshot.summary.actionable_rows == 1
    assert snapshot.summary.critical == 1
    assert snapshot.overall_summary.total_rows == 2

    actionable = next(record for record in snapshot.records if record.should_alert)
    assert actionable.priority == TeamsPriority.CRITICAL
    assert actionable.priority_label.startswith("🔴")
    assert actionable.portal_resources, "portal resources should be populated"
    assert actionable.formatted["billing_amount_display"].startswith("SAR")

    non_alert = next(record for record in snapshot.records if not record.should_alert)
    assert non_alert.priority == TeamsPriority.INFO
    assert non_alert.should_alert is False
    assert non_alert.formatted["status_display"] == "Submitted"