"""
Deep Organizational & Workflow Analyzer for RCM Network Share
Extracts stakeholder channels, workflow branches, organizational structure, and powerful insights
"""

import os
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict, Counter
import json
import re
from typing import Dict, List, Set, Tuple
import openpyxl
from openpyxl import load_workbook


class DeepOrganizationalAnalyzer:
    """Advanced analyzer for extracting organizational insights from RCM rejection data"""
    
    def __init__(self, network_share_path: str):
        self.network_share_path = Path(network_share_path)
        self.insights = {
            'organizational_structure': {},
            'stakeholder_channels': {},
            'workflow_branches': {},
            'temporal_patterns': {},
            'communication_flows': {},
            'process_gaps': {},
            'payer_ecosystems': {},
            'departmental_interactions': {}
        }
        
        # Stakeholder detection patterns
        self.stakeholder_patterns = {
            'departments': ['RCM', 'BILLING', 'CLAIMS', 'AR', 'CODING', 'AUTHORIZATION', 'AUDIT', 'FINANCE'],
            'roles': ['MANAGER', 'COORDINATOR', 'ANALYST', 'SPECIALIST', 'SUPERVISOR', 'DIRECTOR'],
            'payers': ['BUPA', 'TAWUNIYA', 'NCCI', 'MOH', 'ART', 'MALATH', 'SAICO', 'MEDGULF', 'WALAA'],
            'processes': ['INITIAL', 'REJECTION', 'RESUBMISSION', 'APPEAL', 'ESCALATION', 'APPROVAL', 
                         'PENDING', 'REVIEW', 'FOLLOW', 'UPDATE', 'FINAL']
        }
        
        # Workflow stage patterns
        self.workflow_stages = {
            'submission': ['INITIAL', 'SUBMIT', 'NEW', 'UPLOAD'],
            'rejection': ['REJECT', 'DENIED', 'DECLINED', 'REFUSED'],
            'correction': ['CORRECTION', 'FIX', 'UPDATE', 'MODIFY', 'EDIT'],
            'resubmission': ['RESUBMIT', 'RESEND', 'RETRY', 'SECOND'],
            'appeal': ['APPEAL', 'ESCALATE', 'DISPUTE'],
            'approval': ['APPROVED', 'ACCEPTED', 'PAID', 'SETTLED'],
            'follow_up': ['FOLLOW', 'PENDING', 'WAITING', 'HOLD']
        }
    
    def analyze_directory_structure(self) -> Dict:
        """Map the organizational hierarchy from folder structure"""
        print("\n🏢 Analyzing Organizational Structure...")
        
        structure = {
            'root_folders': [],
            'hierarchy_depth': 0,
            'folder_purposes': {},
            'organizational_map': {}
        }
        
        try:
            # Get all directories
            all_dirs = []
            for root, dirs, files in os.walk(self.network_share_path):
                level = root.replace(str(self.network_share_path), '').count(os.sep)
                structure['hierarchy_depth'] = max(structure['hierarchy_depth'], level)
                
                for dir_name in dirs:
                    full_path = Path(root) / dir_name
                    all_dirs.append({
                        'name': dir_name,
                        'path': str(full_path),
                        'level': level + 1,
                        'parent': Path(root).name
                    })
            
            # Analyze root folders
            root_items = list(self.network_share_path.iterdir())
            structure['root_folders'] = [item.name for item in root_items if item.is_dir()]
            
            # Categorize folders by purpose
            for dir_info in all_dirs:
                name = dir_info['name'].upper()
                purposes = []
                
                # Check for payer-specific folders
                for payer in self.stakeholder_patterns['payers']:
                    if payer in name:
                        purposes.append(f'payer_{payer}')
                
                # Check for process-specific folders
                for process in self.stakeholder_patterns['processes']:
                    if process in name:
                        purposes.append(f'process_{process}')
                
                # Check for department folders
                for dept in self.stakeholder_patterns['departments']:
                    if dept in name:
                        purposes.append(f'department_{dept}')
                
                # Check for temporal folders (years, months)
                if re.search(r'20\d{2}', name):  # Year pattern
                    purposes.append('temporal_year')
                if re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)', name):
                    purposes.append('temporal_month')
                
                structure['folder_purposes'][dir_info['name']] = purposes if purposes else ['general']
            
            # Create organizational map
            structure['organizational_map'] = self._create_org_map(all_dirs)
            
        except Exception as e:
            print(f"⚠️ Error analyzing structure: {str(e)}")
        
        self.insights['organizational_structure'] = structure
        return structure
    
    def _create_org_map(self, dirs: List[Dict]) -> Dict:
        """Create hierarchical organizational map"""
        org_map = defaultdict(list)
        
        for dir_info in dirs:
            parent = dir_info['parent']
            child = dir_info['name']
            org_map[parent].append({
                'name': child,
                'level': dir_info['level']
            })
        
        return dict(org_map)
    
    def analyze_stakeholder_channels(self) -> Dict:
        """Identify stakeholder communication channels and interaction patterns"""
        print("\n👥 Analyzing Stakeholder Channels...")
        
        channels = {
            'payer_channels': {},
            'department_channels': {},
            'role_channels': {},
            'cross_functional_teams': [],
            'communication_frequency': {}
        }
        
        try:
            print("   Scanning files (this may take a moment for network shares)...")
            all_files = []
            for root, dirs, files in os.walk(self.network_share_path):
                for file in files:
                    all_files.append(Path(root) / file)
            print(f"   Found {len(all_files)} files")
            
            # Analyze payer channels
            for payer in self.stakeholder_patterns['payers']:
                payer_files = [f for f in all_files if payer in f.name.upper()]
                if payer_files:
                    try:
                        recent_times = []
                        for f in payer_files[:10]:  # Sample for performance
                            try:
                                recent_times.append(f.stat().st_mtime)
                            except:
                                continue
                        recent_activity = max(recent_times) if recent_times else 0
                    except:
                        recent_activity = 0
                    
                    channels['payer_channels'][payer] = {
                        'file_count': len(payer_files),
                        'folders': list(set([f.parent.name for f in payer_files])),
                        'file_types': Counter([f.suffix for f in payer_files]),
                        'recent_activity': recent_activity
                    }
            
            # Analyze department channels
            for dept in self.stakeholder_patterns['departments']:
                dept_files = [f for f in all_files if dept in f.name.upper()]
                if dept_files:
                    channels['department_channels'][dept] = {
                        'file_count': len(dept_files),
                        'associated_payers': self._extract_payers_from_files(dept_files),
                        'activity_level': 'high' if len(dept_files) > 10 else 'medium' if len(dept_files) > 3 else 'low'
                    }
            
            # Identify cross-functional teams (files with multiple stakeholder indicators)
            for f in all_files:
                fname = f.name.upper()
                stakeholders = []
                
                for payer in self.stakeholder_patterns['payers']:
                    if payer in fname:
                        stakeholders.append(f'payer_{payer}')
                
                for dept in self.stakeholder_patterns['departments']:
                    if dept in fname:
                        stakeholders.append(f'dept_{dept}')
                
                if len(stakeholders) >= 2:
                    channels['cross_functional_teams'].append({
                        'file': f.name,
                        'stakeholders': stakeholders,
                        'location': f.parent.name
                    })
            
        except Exception as e:
            print(f"⚠️ Error analyzing stakeholder channels: {str(e)}")
        
        self.insights['stakeholder_channels'] = channels
        return channels
    
    def _extract_payers_from_files(self, files: List[Path]) -> List[str]:
        """Extract payer mentions from file list"""
        payers = set()
        for f in files:
            fname = f.name.upper()
            for payer in self.stakeholder_patterns['payers']:
                if payer in fname:
                    payers.add(payer)
        return list(payers)
    
    def analyze_workflow_branches(self) -> Dict:
        """Map workflow branches and process flows"""
        print("\n🔀 Analyzing Workflow Branches...")
        
        workflows = {
            'process_flows': {},
            'branch_points': [],
            'workflow_stages_identified': {},
            'transition_patterns': [],
            'dead_ends': [],
            'success_paths': []
        }
        
        try:
            all_files = list(self.network_share_path.rglob('*.xlsx'))
            
            # Identify workflow stages from filenames
            for stage_name, keywords in self.workflow_stages.items():
                stage_files = []
                for f in all_files:
                    fname = f.name.upper()
                    if any(kw in fname for kw in keywords):
                        stage_files.append(f.name)
                
                if stage_files:
                    workflows['workflow_stages_identified'][stage_name] = {
                        'file_count': len(stage_files),
                        'files': stage_files[:10],  # Sample
                        'keywords_found': [kw for kw in keywords if any(kw in f.upper() for f in stage_files)]
                    }
            
            # Identify branch points (files indicating decision points)
            branch_indicators = ['VS', 'OR', 'COMPARE', 'REVIEW', 'PENDING', 'HOLD']
            for f in all_files:
                fname = f.name.upper()
                if any(indicator in fname for indicator in branch_indicators):
                    workflows['branch_points'].append({
                        'file': f.name,
                        'type': [ind for ind in branch_indicators if ind in fname][0],
                        'location': f.parent.name
                    })
            
            # Identify transition patterns
            transition_pairs = [
                ('INITIAL', 'REJECTION'),
                ('REJECTION', 'RESUBMISSION'),
                ('RESUBMISSION', 'APPROVAL'),
                ('REJECTION', 'APPEAL'),
                ('PENDING', 'FOLLOW')
            ]
            
            for from_stage, to_stage in transition_pairs:
                files_with_both = [f.name for f in all_files 
                                  if from_stage in f.name.upper() and to_stage in f.name.upper()]
                if files_with_both:
                    workflows['transition_patterns'].append({
                        'from': from_stage,
                        'to': to_stage,
                        'evidence_files': files_with_both,
                        'count': len(files_with_both)
                    })
            
            # Analyze file contents for deeper workflow insights
            workflows['detailed_process_flows'] = self._analyze_excel_workflows(all_files[:10])
            
        except Exception as e:
            print(f"⚠️ Error analyzing workflow branches: {str(e)}")
        
        self.insights['workflow_branches'] = workflows
        return workflows
    
    def _analyze_excel_workflows(self, excel_files: List[Path]) -> Dict:
        """Analyze Excel file contents to identify workflow columns and stages"""
        process_flows = {}
        
        for file_path in excel_files:
            try:
                if not file_path.exists() or not file_path.is_file():
                    continue
                
                wb = load_workbook(file_path, read_only=True, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    # Get headers
                    headers = []
                    for cell in sheet[1]:
                        if cell.value:
                            headers.append(str(cell.value).upper())
                    
                    # Identify workflow-related columns
                    workflow_columns = []
                    status_columns = []
                    
                    for idx, header in enumerate(headers):
                        if any(kw in header for kw in ['STATUS', 'STAGE', 'STEP', 'PHASE']):
                            status_columns.append((idx, header))
                        if any(kw in header for kw in ['WORKFLOW', 'PROCESS', 'ACTION', 'NEXT']):
                            workflow_columns.append((idx, header))
                    
                    if workflow_columns or status_columns:
                        process_flows[f"{file_path.name}_{sheet_name}"] = {
                            'workflow_columns': [col[1] for col in workflow_columns],
                            'status_columns': [col[1] for col in status_columns],
                            'total_columns': len(headers)
                        }
                
                wb.close()
                
            except Exception as e:
                continue
        
        return process_flows
    
    def analyze_temporal_patterns(self) -> Dict:
        """Analyze temporal patterns and time-based insights"""
        print("\n📅 Analyzing Temporal Patterns...")
        
        temporal = {
            'file_age_distribution': {},
            'activity_timeline': {},
            'peak_periods': [],
            'dormant_periods': [],
            'aging_analysis': {},
            'seasonal_patterns': {}
        }
        
        try:
            all_files = [f for f in self.network_share_path.rglob('*') if f.is_file()]
            
            if not all_files:
                return temporal
            
            # Analyze file ages
            now = datetime.now()
            age_buckets = {
                'less_than_month': 0,
                '1_3_months': 0,
                '3_6_months': 0,
                '6_12_months': 0,
                'over_year': 0
            }
            
            modification_dates = []
            
            for f in all_files:
                try:
                    mtime = datetime.fromtimestamp(f.stat().st_mtime)
                    modification_dates.append(mtime)
                    
                    age = (now - mtime).days
                    
                    if age < 30:
                        age_buckets['less_than_month'] += 1
                    elif age < 90:
                        age_buckets['1_3_months'] += 1
                    elif age < 180:
                        age_buckets['3_6_months'] += 1
                    elif age < 365:
                        age_buckets['6_12_months'] += 1
                    else:
                        age_buckets['over_year'] += 1
                
                except:
                    continue
            
            temporal['file_age_distribution'] = age_buckets
            
            # Create activity timeline
            if modification_dates:
                modification_dates.sort()
                
                # Group by month
                monthly_activity = defaultdict(int)
                for dt in modification_dates:
                    month_key = dt.strftime('%Y-%m')
                    monthly_activity[month_key] += 1
                
                temporal['activity_timeline'] = dict(sorted(monthly_activity.items()))
                
                # Identify peak periods (months with >average activity)
                avg_activity = sum(monthly_activity.values()) / len(monthly_activity) if monthly_activity else 0
                temporal['peak_periods'] = [
                    {'month': month, 'file_count': count}
                    for month, count in monthly_activity.items()
                    if count > avg_activity
                ]
                
                # Aging analysis
                oldest_file = min(modification_dates)
                newest_file = max(modification_dates)
                
                temporal['aging_analysis'] = {
                    'oldest_file_date': oldest_file.strftime('%Y-%m-%d'),
                    'newest_file_date': newest_file.strftime('%Y-%m-%d'),
                    'data_span_days': (newest_file - oldest_file).days,
                    'average_file_age_days': int(sum((now - dt).days for dt in modification_dates) / len(modification_dates))
                }
            
        except Exception as e:
            print(f"⚠️ Error analyzing temporal patterns: {str(e)}")
        
        self.insights['temporal_patterns'] = temporal
        return temporal
    
    def analyze_payer_ecosystems(self) -> Dict:
        """Analyze payer-specific ecosystems and their characteristics"""
        print("\n🏥 Analyzing Payer Ecosystems...")
        
        ecosystems = {}
        
        try:
            all_files = list(self.network_share_path.rglob('*.xlsx'))
            
            for payer in self.stakeholder_patterns['payers']:
                payer_files = [f for f in all_files if payer in f.name.upper()]
                
                if not payer_files:
                    continue
                
                ecosystem = {
                    'file_count': len(payer_files),
                    'unique_folders': len(set([f.parent.name for f in payer_files])),
                    'workflow_stages': [],
                    'interaction_points': [],
                    'data_volume_estimate': sum([f.stat().st_size for f in payer_files]) / (1024 * 1024),  # MB
                    'recent_activity': None,
                    'collaboration_patterns': []
                }
                
                # Identify workflow stages for this payer
                for stage_name, keywords in self.workflow_stages.items():
                    stage_files = [f.name for f in payer_files if any(kw in f.name.upper() for kw in keywords)]
                    if stage_files:
                        ecosystem['workflow_stages'].append({
                            'stage': stage_name,
                            'file_count': len(stage_files),
                            'sample_files': stage_files[:3]
                        })
                
                # Identify departments interacting with this payer
                for dept in self.stakeholder_patterns['departments']:
                    dept_payer_files = [f.name for f in payer_files if dept in f.name.upper()]
                    if dept_payer_files:
                        ecosystem['interaction_points'].append({
                            'department': dept,
                            'file_count': len(dept_payer_files)
                        })
                
                # Recent activity
                if payer_files:
                    recent_file = max(payer_files, key=lambda f: f.stat().st_mtime)
                    ecosystem['recent_activity'] = {
                        'file': recent_file.name,
                        'date': datetime.fromtimestamp(recent_file.stat().st_mtime).strftime('%Y-%m-%d')
                    }
                
                # Collaboration patterns (files mentioning multiple entities)
                collab_files = []
                for f in payer_files:
                    fname = f.name.upper()
                    other_payers = [p for p in self.stakeholder_patterns['payers'] if p != payer and p in fname]
                    if other_payers:
                        collab_files.append({
                            'file': f.name,
                            'collaborating_with': other_payers
                        })
                
                ecosystem['collaboration_patterns'] = collab_files
                
                ecosystems[payer] = ecosystem
        
        except Exception as e:
            print(f"⚠️ Error analyzing payer ecosystems: {str(e)}")
        
        self.insights['payer_ecosystems'] = ecosystems
        return ecosystems
    
    def identify_process_gaps(self) -> Dict:
        """Identify missing process steps and potential bottlenecks"""
        print("\n🔍 Identifying Process Gaps...")
        
        gaps = {
            'missing_workflow_stages': [],
            'incomplete_payer_coverage': [],
            'temporal_gaps': [],
            'documentation_gaps': [],
            'recommendations': []
        }
        
        try:
            # Check for missing workflow stages
            expected_stages = list(self.workflow_stages.keys())
            identified_stages = list(self.insights.get('workflow_branches', {}).get('workflow_stages_identified', {}).keys())
            
            missing_stages = [stage for stage in expected_stages if stage not in identified_stages]
            gaps['missing_workflow_stages'] = missing_stages
            
            # Check for payers without complete workflow coverage
            payer_ecosystems = self.insights.get('payer_ecosystems', {})
            
            for payer, ecosystem in payer_ecosystems.items():
                payer_stages = [s['stage'] for s in ecosystem.get('workflow_stages', [])]
                missing_payer_stages = [s for s in expected_stages if s not in payer_stages]
                
                if len(missing_payer_stages) >= 3:  # Missing multiple stages
                    gaps['incomplete_payer_coverage'].append({
                        'payer': payer,
                        'missing_stages': missing_payer_stages,
                        'coverage_percentage': int((len(payer_stages) / len(expected_stages)) * 100)
                    })
            
            # Temporal gaps (periods with no activity)
            activity_timeline = self.insights.get('temporal_patterns', {}).get('activity_timeline', {})
            if activity_timeline:
                months = sorted(activity_timeline.keys())
                for i in range(len(months) - 1):
                    current = datetime.strptime(months[i], '%Y-%m')
                    next_month = datetime.strptime(months[i + 1], '%Y-%m')
                    
                    month_diff = (next_month.year - current.year) * 12 + (next_month.month - current.month)
                    
                    if month_diff > 1:  # Gap detected
                        gaps['temporal_gaps'].append({
                            'from': months[i],
                            'to': months[i + 1],
                            'gap_months': month_diff - 1
                        })
            
            # Generate recommendations
            if missing_stages:
                gaps['recommendations'].append({
                    'priority': 'HIGH',
                    'category': 'workflow_completion',
                    'recommendation': f"Implement missing workflow stages: {', '.join(missing_stages)}",
                    'impact': 'Improve process visibility and tracking'
                })
            
            if gaps['incomplete_payer_coverage']:
                gaps['recommendations'].append({
                    'priority': 'MEDIUM',
                    'category': 'payer_standardization',
                    'recommendation': f"Standardize workflows for {len(gaps['incomplete_payer_coverage'])} payers with incomplete coverage",
                    'impact': 'Ensure consistent processing across all payers'
                })
            
            if gaps['temporal_gaps']:
                gaps['recommendations'].append({
                    'priority': 'LOW',
                    'category': 'data_continuity',
                    'recommendation': f"Investigate {len(gaps['temporal_gaps'])} temporal gaps in data collection",
                    'impact': 'Improve historical data completeness'
                })
        
        except Exception as e:
            print(f"⚠️ Error identifying process gaps: {str(e)}")
        
        self.insights['process_gaps'] = gaps
        return gaps
    
    def generate_deep_insights_report(self, output_path: str = "DEEP_ORGANIZATIONAL_INSIGHTS.md"):
        """Generate comprehensive deep insights report"""
        print("\n📝 Generating Deep Insights Report...")
        
        report = []
        report.append("# Deep Organizational & Workflow Analysis Report")
        report.append(f"\n**Analysis Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append(f"\n**Network Share:** {self.network_share_path}")
        report.append("\n---\n")
        
        # Executive Summary
        report.append("## 🎯 Executive Summary\n")
        org_structure = self.insights.get('organizational_structure', {})
        stakeholders = self.insights.get('stakeholder_channels', {})
        workflows = self.insights.get('workflow_branches', {})
        
        report.append(f"- **Organizational Depth:** {org_structure.get('hierarchy_depth', 0)} levels")
        report.append(f"- **Root Folders:** {len(org_structure.get('root_folders', []))}")
        report.append(f"- **Active Payer Channels:** {len(stakeholders.get('payer_channels', {}))}")
        report.append(f"- **Department Channels:** {len(stakeholders.get('department_channels', {}))}")
        report.append(f"- **Workflow Stages Identified:** {len(workflows.get('workflow_stages_identified', {}))}")
        report.append(f"- **Process Branch Points:** {len(workflows.get('branch_points', []))}")
        report.append("\n")
        
        # Organizational Structure
        report.append("## 🏢 Organizational Structure\n")
        report.append("### Folder Hierarchy\n")
        
        if org_structure.get('root_folders'):
            report.append("**Root Level Folders:**\n")
            for folder in org_structure['root_folders']:
                purposes = org_structure.get('folder_purposes', {}).get(folder, ['unknown'])
                report.append(f"- `{folder}` - Purpose: {', '.join(purposes)}")
            report.append("\n")
        
        # Stakeholder Channels
        report.append("## 👥 Stakeholder Communication Channels\n")
        
        report.append("### Payer Channels\n")
        payer_channels = stakeholders.get('payer_channels', {})
        if payer_channels:
            report.append("| Payer | Files | Folders | Recent Activity |")
            report.append("|-------|-------|---------|-----------------|")
            
            for payer, data in sorted(payer_channels.items(), key=lambda x: x[1]['file_count'], reverse=True):
                recent = datetime.fromtimestamp(data['recent_activity']).strftime('%Y-%m-%d') if data['recent_activity'] else 'N/A'
                report.append(f"| {payer} | {data['file_count']} | {len(data['folders'])} | {recent} |")
            report.append("\n")
        
        report.append("### Department Channels\n")
        dept_channels = stakeholders.get('department_channels', {})
        if dept_channels:
            report.append("| Department | Files | Associated Payers | Activity Level |")
            report.append("|------------|-------|-------------------|----------------|")
            
            for dept, data in sorted(dept_channels.items(), key=lambda x: x[1]['file_count'], reverse=True):
                payers = ', '.join(data['associated_payers']) if data['associated_payers'] else 'None'
                report.append(f"| {dept} | {data['file_count']} | {payers} | {data['activity_level'].upper()} |")
            report.append("\n")
        
        report.append("### Cross-Functional Collaboration\n")
        cross_functional = stakeholders.get('cross_functional_teams', [])
        if cross_functional:
            report.append(f"**Identified {len(cross_functional)} cross-functional collaboration points:**\n")
            for cf in cross_functional[:10]:  # Top 10
                report.append(f"- `{cf['file']}` - Stakeholders: {', '.join(cf['stakeholders'])}")
            report.append("\n")
        
        # Workflow Branches
        report.append("## 🔀 Workflow Branches & Process Flows\n")
        
        report.append("### Identified Workflow Stages\n")
        stages_identified = workflows.get('workflow_stages_identified', {})
        if stages_identified:
            report.append("| Stage | Files | Keywords Found |")
            report.append("|-------|-------|----------------|")
            
            for stage, data in stages_identified.items():
                keywords = ', '.join(data['keywords_found'])
                report.append(f"| {stage.upper()} | {data['file_count']} | {keywords} |")
            report.append("\n")
        
        report.append("### Process Branch Points\n")
        branch_points = workflows.get('branch_points', [])
        if branch_points:
            report.append(f"**Found {len(branch_points)} decision/branch points:**\n")
            
            branch_by_type = defaultdict(list)
            for bp in branch_points:
                branch_by_type[bp['type']].append(bp['file'])
            
            for branch_type, files in branch_by_type.items():
                report.append(f"\n**{branch_type}** ({len(files)} files):")
                for f in files[:5]:
                    report.append(f"- {f}")
            report.append("\n")
        
        report.append("### Workflow Transition Patterns\n")
        transitions = workflows.get('transition_patterns', [])
        if transitions:
            report.append("| From Stage | To Stage | Evidence Files |")
            report.append("|------------|----------|----------------|")
            
            for trans in transitions:
                report.append(f"| {trans['from']} | {trans['to']} | {trans['count']} |")
            report.append("\n")
        
        # Temporal Patterns
        report.append("## 📅 Temporal Patterns & Activity Analysis\n")
        temporal = self.insights.get('temporal_patterns', {})
        
        report.append("### File Age Distribution\n")
        age_dist = temporal.get('file_age_distribution', {})
        if age_dist:
            report.append("| Age Range | File Count |")
            report.append("|-----------|------------|")
            report.append(f"| < 1 month | {age_dist.get('less_than_month', 0)} |")
            report.append(f"| 1-3 months | {age_dist.get('1_3_months', 0)} |")
            report.append(f"| 3-6 months | {age_dist.get('3_6_months', 0)} |")
            report.append(f"| 6-12 months | {age_dist.get('6_12_months', 0)} |")
            report.append(f"| > 1 year | {age_dist.get('over_year', 0)} |")
            report.append("\n")
        
        aging = temporal.get('aging_analysis', {})
        if aging:
            report.append("### Data Timeline\n")
            report.append(f"- **Oldest File:** {aging.get('oldest_file_date', 'N/A')}")
            report.append(f"- **Newest File:** {aging.get('newest_file_date', 'N/A')}")
            report.append(f"- **Data Span:** {aging.get('data_span_days', 0)} days")
            report.append(f"- **Average File Age:** {aging.get('average_file_age_days', 0)} days")
            report.append("\n")
        
        report.append("### Peak Activity Periods\n")
        peaks = temporal.get('peak_periods', [])
        if peaks:
            report.append("| Month | File Count |")
            report.append("|-------|------------|")
            for peak in sorted(peaks, key=lambda x: x['file_count'], reverse=True)[:10]:
                report.append(f"| {peak['month']} | {peak['file_count']} |")
            report.append("\n")
        
        # Payer Ecosystems
        report.append("## 🏥 Payer Ecosystem Analysis\n")
        ecosystems = self.insights.get('payer_ecosystems', {})
        
        for payer, ecosystem in sorted(ecosystems.items(), key=lambda x: x[1]['file_count'], reverse=True):
            report.append(f"### {payer}\n")
            report.append(f"- **Files:** {ecosystem['file_count']}")
            report.append(f"- **Folders:** {ecosystem['unique_folders']}")
            report.append(f"- **Data Volume:** {ecosystem['data_volume_estimate']:.2f} MB")
            
            if ecosystem.get('recent_activity'):
                report.append(f"- **Recent Activity:** {ecosystem['recent_activity']['file']} ({ecosystem['recent_activity']['date']})")
            
            report.append("\n**Workflow Stages:**")
            for stage in ecosystem.get('workflow_stages', []):
                report.append(f"- {stage['stage'].upper()}: {stage['file_count']} files")
            
            report.append("\n**Department Interactions:**")
            for interaction in ecosystem.get('interaction_points', []):
                report.append(f"- {interaction['department']}: {interaction['file_count']} touchpoints")
            
            if ecosystem.get('collaboration_patterns'):
                report.append(f"\n**Cross-Payer Collaborations:** {len(ecosystem['collaboration_patterns'])} instances")
            
            report.append("\n")
        
        # Process Gaps
        report.append("## 🔍 Process Gaps & Recommendations\n")
        gaps = self.insights.get('process_gaps', {})
        
        if gaps.get('missing_workflow_stages'):
            report.append("### ⚠️ Missing Workflow Stages\n")
            for stage in gaps['missing_workflow_stages']:
                report.append(f"- {stage.upper()}")
            report.append("\n")
        
        if gaps.get('incomplete_payer_coverage'):
            report.append("### 📊 Payers with Incomplete Coverage\n")
            report.append("| Payer | Missing Stages | Coverage % |")
            report.append("|-------|----------------|------------|")
            for item in gaps['incomplete_payer_coverage']:
                missing = ', '.join(item['missing_stages'])
                report.append(f"| {item['payer']} | {len(item['missing_stages'])} | {item['coverage_percentage']}% |")
            report.append("\n")
        
        if gaps.get('temporal_gaps'):
            report.append("### 📅 Temporal Data Gaps\n")
            for gap in gaps['temporal_gaps']:
                report.append(f"- Gap from {gap['from']} to {gap['to']} ({gap['gap_months']} months)")
            report.append("\n")
        
        report.append("### 💡 Strategic Recommendations\n")
        for idx, rec in enumerate(gaps.get('recommendations', []), 1):
            report.append(f"\n**{idx}. {rec['recommendation']}**")
            report.append(f"- **Priority:** {rec['priority']}")
            report.append(f"- **Category:** {rec['category']}")
            report.append(f"- **Expected Impact:** {rec['impact']}")
        
        report.append("\n")
        
        # Key Insights
        report.append("## 🎯 Key Strategic Insights\n")
        report.append(self._generate_strategic_insights())
        
        # Save report
        output_file = Path(output_path)
        output_file.write_text('\n'.join(report), encoding='utf-8')
        print(f"✅ Deep insights report saved to: {output_path}")
        
        return output_path
    
    def _generate_strategic_insights(self) -> str:
        """Generate strategic insights based on all analysis"""
        insights = []
        
        # Analyze stakeholder balance
        payer_channels = self.insights.get('stakeholder_channels', {}).get('payer_channels', {})
        if payer_channels:
            file_counts = [data['file_count'] for data in payer_channels.values()]
            max_files = max(file_counts) if file_counts else 0
            min_files = min(file_counts) if file_counts else 0
            
            if max_files > min_files * 3:
                insights.append(f"1. **Stakeholder Imbalance:** High variation in payer engagement ({max_files} vs {min_files} files). Consider standardizing data collection across all payers.")
        
        # Analyze workflow completeness
        workflows = self.insights.get('workflow_branches', {}).get('workflow_stages_identified', {})
        if len(workflows) < 5:
            insights.append(f"2. **Workflow Coverage Gap:** Only {len(workflows)} workflow stages identified. Comprehensive tracking requires visibility across all stages (submission → rejection → correction → resubmission → approval).")
        
        # Analyze temporal patterns
        temporal = self.insights.get('temporal_patterns', {})
        aging = temporal.get('aging_analysis', {})
        if aging and aging.get('average_file_age_days', 0) > 90:
            insights.append(f"3. **Data Freshness Concern:** Average file age is {aging['average_file_age_days']} days. Consider implementing automated real-time data collection.")
        
        # Analyze collaboration
        cross_functional = self.insights.get('stakeholder_channels', {}).get('cross_functional_teams', [])
        if len(cross_functional) < 5:
            insights.append("4. **Limited Cross-Functional Collaboration:** Few files show multi-stakeholder involvement. Enhance communication between departments for better claim resolution.")
        
        # Analyze payer ecosystems
        ecosystems = self.insights.get('payer_ecosystems', {})
        payers_with_limited_stages = sum(1 for eco in ecosystems.values() if len(eco.get('workflow_stages', [])) < 3)
        if payers_with_limited_stages > 0:
            insights.append(f"5. **Process Standardization Needed:** {payers_with_limited_stages} payers have limited workflow visibility. Implement standardized tracking across all payer relationships.")
        
        # Analyze branch points
        branch_points = self.insights.get('workflow_branches', {}).get('branch_points', [])
        if len(branch_points) > 10:
            insights.append(f"6. **High Decision Complexity:** {len(branch_points)} decision/branch points detected. Consider workflow automation to reduce manual decision-making and improve efficiency.")
        
        if not insights:
            insights.append("1. **Well-Structured Organization:** Data organization appears comprehensive with good stakeholder representation and workflow coverage.")
        
        return '\n'.join(insights) + "\n"
    
    def save_insights_json(self, output_path: str = "DEEP_ORGANIZATIONAL_INSIGHTS.json"):
        """Save insights as JSON"""
        print(f"\n💾 Saving insights to JSON: {output_path}")
        
        # Convert any non-serializable objects
        serializable_insights = json.loads(
            json.dumps(self.insights, default=str)
        )
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(serializable_insights, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Insights saved to: {output_path}")
        return output_path
    
    def run_full_analysis(self):
        """Run complete deep analysis"""
        print("="*80)
        print("🚀 DEEP ORGANIZATIONAL & WORKFLOW ANALYSIS")
        print("="*80)
        
        # Run all analyses
        self.analyze_directory_structure()
        self.analyze_stakeholder_channels()
        self.analyze_workflow_branches()
        self.analyze_temporal_patterns()
        self.analyze_payer_ecosystems()
        self.identify_process_gaps()
        
        # Generate outputs
        self.generate_deep_insights_report()
        self.save_insights_json()
        
        print("\n" + "="*80)
        print("✅ ANALYSIS COMPLETE")
        print("="*80)
        
        # Summary
        print("\n📊 ANALYSIS SUMMARY:")
        print(f"   - Organizational Depth: {self.insights['organizational_structure'].get('hierarchy_depth', 0)} levels")
        print(f"   - Payer Channels: {len(self.insights['stakeholder_channels'].get('payer_channels', {}))}")
        print(f"   - Department Channels: {len(self.insights['stakeholder_channels'].get('department_channels', {}))}")
        print(f"   - Workflow Stages: {len(self.insights['workflow_branches'].get('workflow_stages_identified', {}))}")
        print(f"   - Branch Points: {len(self.insights['workflow_branches'].get('branch_points', []))}")
        print(f"   - Process Gaps Identified: {len(self.insights['process_gaps'].get('recommendations', []))}")
        
        print("\n📁 Generated Files:")
        print("   - DEEP_ORGANIZATIONAL_INSIGHTS.md")
        print("   - DEEP_ORGANIZATIONAL_INSIGHTS.json")
        print()


if __name__ == "__main__":
    # Network share path
    network_share = r"\\128.1.1.86\InmaRCMRejection"
    
    print("Initializing Deep Organizational Analyzer...")
    analyzer = DeepOrganizationalAnalyzer(network_share)
    
    # Run full analysis
    analyzer.run_full_analysis()
    
    print("\n🎯 Deep analysis complete! Check the generated files for comprehensive insights.")
