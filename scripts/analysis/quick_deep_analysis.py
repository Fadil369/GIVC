"""
Quick Deep Analysis - Uses local analysis_data folder for fast insights
Extracts organizational structure, stakeholder patterns, and workflow insights
"""

import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import json
import re
import openpyxl
from openpyxl import load_workbook


def analyze_network_share_structure():
    """Analyze the network share structure with detailed insights"""
    
    network_share = r"\\128.1.1.86\InmaRCMRejection"
    
    insights = {
        'metadata': {
            'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'network_share': network_share
        },
        'organizational_structure': {},
        'stakeholder_analysis': {},
        'workflow_intelligence': {},
        'file_analytics': {},
        'strategic_insights': []
    }
    
    print("="*80)
    print("🔍 DEEP ORGANIZATIONAL INTELLIGENCE ANALYSIS")
    print("="*80)
    
    # Phase 1: Directory Structure Analysis
    print("\n📂 Phase 1: Organizational Structure Analysis")
    print("-" * 60)
    
    try:
        root_items = list(Path(network_share).iterdir())
        folders = [item for item in root_items if item.is_dir()]
        files = [item for item in root_items if item.is_file()]
        
        insights['organizational_structure'] = {
            'total_root_items': len(root_items),
            'folders': len(folders),
            'files': len(files),
            'folder_names': [f.name for f in folders],
            'file_names': [f.name for f in files]
        }
        
        print(f"   ✓ Root folders: {len(folders)}")
        print(f"   ✓ Root files: {len(files)}")
        
        # Categorize folders
        payer_folders = []
        process_folders = []
        temporal_folders = []
        department_folders = []
        
        payer_keywords = ['BUPA', 'TAWUNIYA', 'NCCI', 'MOH', 'ART', 'MALATH', 'SAICO', 'MEDGULF', 'WALAA']
        process_keywords = ['REJECTION', 'RESUBMISSION', 'INITIAL', 'APPROVAL', 'PENDING', 'REVIEW']
        dept_keywords = ['RCM', 'BILLING', 'CLAIMS', 'AR', 'CODING']
        
        for folder in folders:
            fname = folder.name.upper()
            
            if any(p in fname for p in payer_keywords):
                payer_folders.append(folder.name)
            
            if any(p in fname for p in process_keywords):
                process_folders.append(folder.name)
            
            if re.search(r'20\d{2}', fname):
                temporal_folders.append(folder.name)
            
            if any(d in fname for d in dept_keywords):
                department_folders.append(folder.name)
        
        insights['organizational_structure']['categorization'] = {
            'payer_specific': payer_folders,
            'process_specific': process_folders,
            'temporal': temporal_folders,
            'department_specific': department_folders
        }
        
        print(f"   ✓ Payer-specific folders: {len(payer_folders)}")
        print(f"   ✓ Process-specific folders: {len(process_folders)}")
        print(f"   ✓ Temporal folders: {len(temporal_folders)}")
        print(f"   ✓ Department folders: {len(department_folders)}")
        
    except Exception as e:
        print(f"   ⚠️ Error accessing network share: {str(e)}")
        insights['organizational_structure']['error'] = str(e)
    
    # Phase 2: File Analysis from local analysis_data
    print("\n📊 Phase 2: File Intelligence Analysis")
    print("-" * 60)
    
    analysis_data_path = Path("analysis_data")
    
    if analysis_data_path.exists():
        excel_files = list(analysis_data_path.glob("*.xlsx"))
        print(f"   Analyzing {len(excel_files)} Excel files...")
        
        file_analytics = {
            'total_files': len(excel_files),
            'files_analyzed': [],
            'naming_patterns': {},
            'content_insights': {}
        }
        
        # Analyze file naming patterns
        naming_elements = {
            'payers': [],
            'processes': [],
            'time_indicators': [],
            'versions': []
        }
        
        payer_keywords = ['BUPA', 'TAWUNIYA', 'NCCI', 'MOH', 'ART', 'MALATH', 'SAICO']
        process_keywords = ['REJECTION', 'RESUBMISSION', 'INITIAL', 'REJ', 'APPROVAL', 'UPDATE']
        
        for file in excel_files:
            fname = file.name.upper()
            file_info = {
                'name': file.name,
                'size_mb': file.stat().st_size / (1024 * 1024),
                'modified': datetime.fromtimestamp(file.stat().st_mtime).strftime('%Y-%m-%d')
            }
            
            # Extract naming elements
            for payer in payer_keywords:
                if payer in fname:
                    naming_elements['payers'].append(payer)
                    file_info['payer'] = payer
            
            for process in process_keywords:
                if process in fname:
                    naming_elements['processes'].append(process)
                    file_info['process'] = process
            
            # Check for version indicators
            if 'COPY' in fname or 'V1' in fname or 'V2' in fname:
                naming_elements['versions'].append(file.name)
                file_info['is_version'] = True
            
            # Check for time indicators
            years = re.findall(r'20\d{2}', fname)
            if years:
                naming_elements['time_indicators'].extend(years)
                file_info['year_indicators'] = years
            
            file_analytics['files_analyzed'].append(file_info)
        
        file_analytics['naming_patterns'] = {
            'payer_mentions': dict(Counter(naming_elements['payers'])),
            'process_mentions': dict(Counter(naming_elements['processes'])),
            'versioned_files': len(naming_elements['versions']),
            'temporal_indicators': dict(Counter(naming_elements['time_indicators']))
        }
        
        insights['file_analytics'] = file_analytics
        
        print(f"   ✓ Files analyzed: {len(excel_files)}")
        print(f"   ✓ Payer mentions: {dict(Counter(naming_elements['payers']))}")
        print(f"   ✓ Process indicators: {dict(Counter(naming_elements['processes']))}")
    
    # Phase 3: Deep Content Analysis
    print("\n🔬 Phase 3: Deep Content Intelligence")
    print("-" * 60)
    
    stakeholder_analysis = {
        'payer_ecosystems': {},
        'workflow_patterns': {},
        'communication_channels': {},
        'process_maturity': {}
    }
    
    if analysis_data_path.exists():
        excel_files = list(analysis_data_path.glob("*.xlsx"))
        
        for file in excel_files:
            try:
                wb = load_workbook(file, read_only=True, data_only=True)
                
                file_insights = {
                    'sheets': [],
                    'total_sheets': len(wb.sheetnames),
                    'data_volume_rows': 0,
                    'key_columns': [],
                    'workflow_indicators': []
                }
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    # Get headers
                    headers = []
                    for cell in sheet[1]:
                        if cell.value:
                            headers.append(str(cell.value).upper())
                    
                    # Count rows
                    row_count = sum(1 for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row))
                    file_insights['data_volume_rows'] += row_count
                    
                    # Identify key columns
                    workflow_columns = []
                    stakeholder_columns = []
                    financial_columns = []
                    temporal_columns = []
                    
                    for header in headers:
                        if any(kw in header for kw in ['STATUS', 'STATE', 'STAGE', 'WORKFLOW', 'PROCESS']):
                            workflow_columns.append(header)
                        
                        if any(kw in header for kw in ['PAYER', 'INSURANCE', 'COMPANY', 'PROVIDER']):
                            stakeholder_columns.append(header)
                        
                        if any(kw in header for kw in ['AMOUNT', 'VALUE', 'SAR', 'PRICE', 'COST', 'TOTAL']):
                            financial_columns.append(header)
                        
                        if any(kw in header for kw in ['DATE', 'TIME', 'CREATED', 'MODIFIED', 'SUBMITTED']):
                            temporal_columns.append(header)
                    
                    sheet_info = {
                        'name': sheet_name,
                        'rows': row_count,
                        'columns': len(headers),
                        'workflow_columns': workflow_columns,
                        'stakeholder_columns': stakeholder_columns,
                        'financial_columns': financial_columns,
                        'temporal_columns': temporal_columns
                    }
                    
                    file_insights['sheets'].append(sheet_info)
                    file_insights['key_columns'].extend(workflow_columns + stakeholder_columns)
                    
                    if workflow_columns:
                        file_insights['workflow_indicators'].extend(workflow_columns)
                
                stakeholder_analysis['payer_ecosystems'][file.name] = file_insights
                
                print(f"   ✓ {file.name}: {file_insights['data_volume_rows']} rows, {file_insights['total_sheets']} sheets")
                
                wb.close()
                
            except Exception as e:
                print(f"   ⚠️ Error analyzing {file.name}: {str(e)}")
    
    insights['stakeholder_analysis'] = stakeholder_analysis
    
    # Phase 4: Workflow Intelligence
    print("\n🔀 Phase 4: Workflow & Process Intelligence")
    print("-" * 60)
    
    workflow_intelligence = {
        'identified_processes': [],
        'workflow_stages': {},
        'transition_patterns': [],
        'bottleneck_indicators': []
    }
    
    # Identify workflow stages from file names and content
    workflow_stages = {
        'submission': 0,
        'rejection': 0,
        'correction': 0,
        'resubmission': 0,
        'approval': 0,
        'pending': 0
    }
    
    stage_keywords = {
        'submission': ['INITIAL', 'SUBMIT', 'NEW'],
        'rejection': ['REJECT', 'DENIED', 'REJ'],
        'correction': ['CORRECTION', 'FIX', 'UPDATE'],
        'resubmission': ['RESUBMIT', 'RESEND', 'RETRY'],
        'approval': ['APPROVED', 'ACCEPTED', 'PAID'],
        'pending': ['PENDING', 'HOLD', 'WAITING']
    }
    
    if analysis_data_path.exists():
        all_files = list(analysis_data_path.glob("*.xlsx"))
        
        for file in all_files:
            fname = file.name.upper()
            
            for stage, keywords in stage_keywords.items():
                if any(kw in fname for kw in keywords):
                    workflow_stages[stage] += 1
    
    workflow_intelligence['workflow_stages'] = workflow_stages
    workflow_intelligence['identified_processes'] = [
        stage for stage, count in workflow_stages.items() if count > 0
    ]
    
    # Identify transition patterns
    transition_indicators = []
    if analysis_data_path.exists():
        for file in all_files:
            fname = file.name.upper()
            
            # Look for transition indicators (e.g., "REJECTION TO RESUBMISSION", "INITIAL VS FINAL")
            if 'VS' in fname:
                transition_indicators.append({
                    'file': file.name,
                    'type': 'comparison',
                    'pattern': 'A vs B'
                })
            
            if any(word in fname for word in ['FOLLOW', 'SECOND', 'RETRY']):
                transition_indicators.append({
                    'file': file.name,
                    'type': 'iterative_process',
                    'pattern': 'repeated_attempts'
                })
    
    workflow_intelligence['transition_patterns'] = transition_indicators
    
    print(f"   ✓ Workflow stages identified: {len(workflow_intelligence['identified_processes'])}")
    print(f"   ✓ Stages: {', '.join(workflow_intelligence['identified_processes'])}")
    print(f"   ✓ Transition patterns: {len(transition_indicators)}")
    
    insights['workflow_intelligence'] = workflow_intelligence
    
    # Phase 5: Strategic Insights
    print("\n🎯 Phase 5: Strategic Intelligence & Recommendations")
    print("-" * 60)
    
    strategic_insights = []
    
    # Analyze payer coverage
    payer_mentions = file_analytics.get('naming_patterns', {}).get('payer_mentions', {})
    if payer_mentions:
        dominant_payer = max(payer_mentions, key=payer_mentions.get)
        strategic_insights.append({
            'insight': f"Payer Focus Imbalance",
            'details': f"{dominant_payer} has {payer_mentions[dominant_payer]} file mentions, indicating high activity or complexity",
            'recommendation': f"Investigate {dominant_payer} rejection patterns for systemic issues",
            'priority': 'HIGH',
            'impact': 'Process optimization opportunity'
        })
    
    # Analyze workflow completeness
    identified_stages = len(workflow_intelligence['identified_processes'])
    if identified_stages < 5:
        strategic_insights.append({
            'insight': "Incomplete Workflow Visibility",
            'details': f"Only {identified_stages}/6 workflow stages have documentation",
            'recommendation': "Implement comprehensive tracking across all claim lifecycle stages",
            'priority': 'MEDIUM',
            'impact': 'Improved process control and metrics'
        })
    
    # Analyze versioning
    versioned_files = file_analytics.get('naming_patterns', {}).get('versioned_files', 0)
    if versioned_files > 2:
        strategic_insights.append({
            'insight': "Manual Version Control",
            'details': f"{versioned_files} files with version indicators (Copy, V1, V2)",
            'recommendation': "Implement automated version control and single source of truth",
            'priority': 'MEDIUM',
            'impact': 'Reduce errors and confusion'
        })
    
    # Analyze transition patterns
    if len(transition_indicators) > 0:
        strategic_insights.append({
            'insight': "Active Process Refinement",
            'details': f"{len(transition_indicators)} files show comparison/iteration patterns",
            'recommendation': "Analyze these transition patterns to identify improvement opportunities",
            'priority': 'HIGH',
            'impact': 'Reduce resubmission cycles and improve first-time success rate'
        })
    
    # Data organization insight
    total_sheets = sum(
        eco.get('total_sheets', 0) 
        for eco in stakeholder_analysis.get('payer_ecosystems', {}).values()
    )
    
    if total_sheets > 20:
        strategic_insights.append({
            'insight': "High Data Fragmentation",
            'details': f"{total_sheets} sheets across {len(excel_files)} files",
            'recommendation': "Consolidate data into centralized database with unified schema",
            'priority': 'HIGH',
            'impact': 'Faster analysis, automated reporting, real-time insights'
        })
    
    insights['strategic_insights'] = strategic_insights
    
    print(f"   ✓ Strategic insights generated: {len(strategic_insights)}")
    for idx, insight in enumerate(strategic_insights, 1):
        print(f"   {idx}. [{insight['priority']}] {insight['insight']}")
    
    # Save insights
    print("\n💾 Saving Analysis Results...")
    print("-" * 60)
    
    # Save JSON
    json_path = "DEEP_ORGANIZATIONAL_INSIGHTS.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(insights, f, indent=2, ensure_ascii=False)
    print(f"   ✓ JSON saved: {json_path}")
    
    # Generate Markdown Report
    md_path = "DEEP_ORGANIZATIONAL_INSIGHTS.md"
    generate_markdown_report(insights, md_path)
    print(f"   ✓ Report saved: {md_path}")
    
    print("\n" + "="*80)
    print("✅ DEEP ANALYSIS COMPLETE")
    print("="*80)
    
    return insights


def generate_markdown_report(insights, output_path):
    """Generate comprehensive markdown report"""
    
    report = []
    report.append("# Deep Organizational Intelligence Analysis")
    report.append(f"\n**Analysis Date:** {insights['metadata']['analysis_date']}")
    report.append(f"**Network Share:** `{insights['metadata']['network_share']}`")
    report.append("\n---\n")
    
    # Executive Summary
    report.append("## 🎯 Executive Summary\n")
    
    org = insights.get('organizational_structure', {})
    analytics = insights.get('file_analytics', {})
    workflows = insights.get('workflow_intelligence', {})
    
    report.append("### Key Metrics\n")
    report.append(f"- **Root Folders:** {org.get('folders', 0)}")
    report.append(f"- **Files Analyzed:** {analytics.get('total_files', 0)}")
    report.append(f"- **Workflow Stages Identified:** {len(workflows.get('identified_processes', []))}")
    report.append(f"- **Strategic Insights:** {len(insights.get('strategic_insights', []))}")
    report.append("\n")
    
    # Organizational Structure
    report.append("## 🏢 Organizational Structure\n")
    
    categorization = org.get('categorization', {})
    
    if categorization.get('payer_specific'):
        report.append("### Payer-Specific Folders\n")
        for folder in categorization['payer_specific']:
            report.append(f"- `{folder}`")
        report.append("\n")
    
    if categorization.get('process_specific'):
        report.append("### Process-Specific Folders\n")
        for folder in categorization['process_specific']:
            report.append(f"- `{folder}`")
        report.append("\n")
    
    if categorization.get('temporal'):
        report.append("### Temporal Organization\n")
        for folder in categorization['temporal']:
            report.append(f"- `{folder}`")
        report.append("\n")
    
    # File Analytics
    report.append("## 📊 File Intelligence Analysis\n")
    
    naming_patterns = analytics.get('naming_patterns', {})
    
    report.append("### Payer Distribution in Files\n")
    payer_mentions = naming_patterns.get('payer_mentions', {})
    if payer_mentions:
        report.append("| Payer | Mentions |")
        report.append("|-------|----------|")
        for payer, count in sorted(payer_mentions.items(), key=lambda x: x[1], reverse=True):
            report.append(f"| {payer} | {count} |")
        report.append("\n")
    
    report.append("### Process Keywords in Files\n")
    process_mentions = naming_patterns.get('process_mentions', {})
    if process_mentions:
        report.append("| Process | Mentions |")
        report.append("|---------|----------|")
        for process, count in sorted(process_mentions.items(), key=lambda x: x[1], reverse=True):
            report.append(f"| {process} | {count} |")
        report.append("\n")
    
    report.append(f"**Versioned Files:** {naming_patterns.get('versioned_files', 0)}\n\n")
    
    # Stakeholder Analysis
    report.append("## 👥 Stakeholder & Content Analysis\n")
    
    payer_ecosystems = insights.get('stakeholder_analysis', {}).get('payer_ecosystems', {})
    
    if payer_ecosystems:
        report.append("| File | Sheets | Rows | Workflow Columns | Financial Columns |")
        report.append("|------|--------|------|------------------|-------------------|")
        
        for file_name, ecosystem in payer_ecosystems.items():
            total_workflow = sum(len(sheet.get('workflow_columns', [])) for sheet in ecosystem.get('sheets', []))
            total_financial = sum(len(sheet.get('financial_columns', [])) for sheet in ecosystem.get('sheets', []))
            
            report.append(f"| {file_name} | {ecosystem.get('total_sheets', 0)} | "
                         f"{ecosystem.get('data_volume_rows', 0)} | {total_workflow} | {total_financial} |")
        report.append("\n")
    
    # Workflow Intelligence
    report.append("## 🔀 Workflow & Process Intelligence\n")
    
    report.append("### Identified Workflow Stages\n")
    workflow_stages = workflows.get('workflow_stages', {})
    if workflow_stages:
        report.append("| Stage | Evidence Count |")
        report.append("|-------|----------------|")
        for stage, count in sorted(workflow_stages.items(), key=lambda x: x[1], reverse=True):
            if count > 0:
                report.append(f"| {stage.upper()} | {count} |")
        report.append("\n")
    
    report.append("### Transition Patterns\n")
    transitions = workflows.get('transition_patterns', [])
    if transitions:
        for trans in transitions:
            report.append(f"- **{trans['type']}**: `{trans['file']}` ({trans['pattern']})")
        report.append("\n")
    else:
        report.append("No explicit transition patterns detected in file names.\n\n")
    
    # Strategic Insights
    report.append("## 🎯 Strategic Insights & Recommendations\n")
    
    strategic = insights.get('strategic_insights', [])
    
    for idx, insight in enumerate(strategic, 1):
        report.append(f"### {idx}. {insight['insight']} [{insight['priority']}]\n")
        report.append(f"**Details:** {insight['details']}\n\n")
        report.append(f"**Recommendation:** {insight['recommendation']}\n\n")
        report.append(f"**Expected Impact:** {insight['impact']}\n\n")
    
    # Integration Roadmap
    report.append("## 🚀 Integration Roadmap\n")
    report.append("### Phase 1: Data Consolidation (Immediate)\n")
    report.append("1. Create unified database schema for all rejection data\n")
    report.append("2. Implement automated ETL pipeline from network share to database\n")
    report.append("3. Establish single source of truth for claim lifecycle tracking\n\n")
    
    report.append("### Phase 2: Process Automation (1-2 weeks)\n")
    report.append("1. Integrate resubmission service with identified workflow patterns\n")
    report.append("2. Implement payer-specific validation rules\n")
    report.append("3. Deploy automated rejection analysis and correction\n\n")
    
    report.append("### Phase 3: Intelligence Layer (2-4 weeks)\n")
    report.append("1. Build real-time analytics dashboard\n")
    report.append("2. Implement predictive rejection prevention\n")
    report.append("3. Deploy stakeholder-specific portals\n\n")
    
    # Conclusion
    report.append("## 📝 Conclusion\n")
    report.append(f"This deep analysis identified **{len(strategic)} strategic opportunities** for ")
    report.append("process improvement and automation. The organizational structure reveals ")
    report.append(f"**{org.get('folders', 0)} primary channels** with ")
    report.append(f"**{len(workflows.get('identified_processes', []))} active workflow stages**.\n\n")
    report.append("**Next Steps:**\n")
    report.append("1. Review strategic insights with stakeholders\n")
    report.append("2. Prioritize implementations based on ROI analysis\n")
    report.append("3. Begin Phase 1 data consolidation\n")
    report.append("4. Schedule weekly progress reviews\n")
    
    # Save report
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report))


if __name__ == "__main__":
    analyze_network_share_structure()
